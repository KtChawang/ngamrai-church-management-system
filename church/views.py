
 # views.py
from .models import ChurchGalleryImage 
from django.db.models import Sum, F, FloatField
from django.http import FileResponse
from calendar import month_name
from django.db.models.functions import ExtractMonth, ExtractYear
from .models import ChurchStaffMember, StaffPayment
from .forms import ChurchStaffMemberForm, StaffPaymentForm
from .forms import ChurchOnlineGivingForm
from django.http import HttpResponseRedirect, JsonResponse
from django.contrib.auth.decorators import permission_required
from .utils import get_online_members
from django.views.decorators.http import require_GET, require_POST
from .models import MemberChatMessage
from django.utils.timesince import timesince
from .models import Subscription
from django.utils import timezone
from .models import PaymentQRCode
from .forms import LicenseValidationForm
from .models import LicenseKey
from .models import ChairmanMessage
from .models import Testimonial
from .forms import AdvertisementForm
from .models import Advertisement
from church.forms import VolunteerForm
from django.core.exceptions import PermissionDenied
from .models import PrayerRequest
from .forms import PrayerRequestForm
from .utils import get_scripture_passage
from .models import ActivityLog
from django.http import JsonResponse, HttpResponseBadRequest
from .forms import BibleReadingPlanForm
from .models import BibleReadingPlan, MemberReadingLog
from .models import SMSLog
from .utils import send_sms
from xhtml2pdf import pisa
from church.utils import send_sms
from itertools import chain, groupby
from .forms import OnlineGivingForm
from .models import OnlineGiving
from django.http import HttpResponseForbidden
from .mixins import IsChurchMemberMixin
from django.views.generic import TemplateView
from django.views.generic.edit import FormView
from django.views.decorators.csrf import csrf_exempt
from django.views.generic import ListView, DetailView, View
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from .models import Notification
from django.http import JsonResponse
from .models import Event
from .forms import EventForm
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from urllib.parse import urlencode
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from datetime import timedelta
from itertools import groupby
from django.contrib.contenttypes.models import ContentType
from .models import StewardshipRecord, AuditLog
from django.db import models
from django.db.models import Sum
from .forms import SundayReceiptFilterForm
from django.views.decorators.http import require_POST
from decimal import Decimal
from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import get_object_or_404, render, redirect
from .models import SundayIncomeReceipt
from .forms import SundayIncomeReceiptForm
from django.template.loader import render_to_string
from django.shortcuts import get_object_or_404
from django.conf import settings
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from django.templatetags.static import static
from django.utils.html import format_html
from urllib.parse import quote
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import xlsxwriter
from .utils import render_pdf
import io
from django.utils.dateparse import parse_date
import pandas as pd
from collections import OrderedDict, defaultdict
import openpyxl
from datetime import datetime, date
from django.db.models.functions import TruncMonth
from django.utils.dateformat import DateFormat
from django.utils.formats import date_format
import calendar
import json
from django.contrib.auth import login as auth_login
from django.contrib.auth.views import LoginView, LogoutView
from django.urls import reverse_lazy
from django.http import HttpResponse, HttpResponseNotFound
from django.middleware.csrf import get_token
from django.views.decorators.csrf import csrf_protect
from django.contrib.sessions.backends.db import SessionStore
from django.views.decorators.http import require_http_methods
from django.db import transaction
from django.db import IntegrityError
from .models import SEX_CHOICES, MARITAL_STATUS_CHOICES
import os
import tempfile
from io import BytesIO
from django import forms
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.utils.timezone import now
from django.contrib.auth import get_user_model, authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.hashers import make_password, check_password
from django.core.exceptions import ObjectDoesNotExist
from django.db.models import Sum, Q, Count, Prefetch
from django.db.models.signals import post_save
from django.dispatch import receiver
from django.http import HttpResponse, Http404
from django.template.loader import get_template
from django.urls import reverse, reverse_lazy
from django.utils.timezone import localtime
from django.utils.decorators import method_decorator, decorator_from_middleware
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from weasyprint import HTML, CSS
from .forms import ChatRoomForm

from .middleware import ChurchAdminSessionMiddleware  # ✅ correct
from .models import (
    Church, ChurchAdminProfile, CustomUser, Event, Announcement, LiveStream,
    Member, MemberCategory, Volunteer, StewardshipRecord, ChatRoom, 
)
from .forms import (
    StewardshipForm, StewardshipRecordForm, ChurchRegistrationForm,
    ChurchAdminLoginForm, ChurchImageForm, MemberRegistrationForm,
    VolunteerCreationForm,
)



User = get_user_model()

def register_church(request):
    """
    Church sign‑up that validates a licence key, creates the church,
    and guarantees there is ONE (and only one) Subscription.
    """
    if request.method == 'POST':
        form = ChurchRegistrationForm(request.POST)
        if form.is_valid():
            key_input  = form.cleaned_data['license_key'].strip()
            email      = form.cleaned_data['email_address']
            password   = form.cleaned_data['password']
            full_name  = form.cleaned_data['church_name']

            # ── 1. Licence‑key sanity check ───────────────────────────
            try:
                lic = LicenseKey.objects.get(key=key_input)
                if lic.is_active or lic.is_expired():
                    form.add_error('license_key',
                                   "This licence key is already used or expired.")
                    return render(request,
                                  'church/register_church.html',
                                  {'form': form})
            except LicenseKey.DoesNotExist:
                form.add_error('license_key', "Invalid licence key.")
                return render(request,
                              'church/register_church.html',
                              {'form': form})

            # ── 2. Duplicate‑email guard (cheap check) ───────────────
            if CustomUser.objects.filter(email=email).exists():
                form.add_error('email_address',
                               "A user with this email already exists.")
                return render(request,
                              'church/register_church.html',
                              {'form': form})

            # ── 3. Atomic create block  ──────────────────────────────
            try:
                with transaction.atomic():
                    # 3‑a: user
                    user = CustomUser.objects.create_user(
                        email=email,
                        password=password,
                        is_church_admin=True,
                    )

                    # 3‑b: church
                    church = form.save(commit=False)
                    church.email_address = email
                    church.admin = user
                    church.save()

                    # 3‑c: ONE subscription (create or reuse)
                    Subscription.objects.get_or_create(
                        church=church,
                        defaults={
                            "plan": "monthly",
                            "amount": 0,
                            "last_paid_at": timezone.now().date(),
                        },
                    )

                    # 3‑d: admin profile
                    ChurchAdminProfile.objects.get_or_create(
                        church=church,
                        defaults={
                            'user': user,
                            'full_name': full_name,
                        }
                    )

                    # 3‑e: bind the licence key
                    lic.is_active       = True
                    lic.church_name     = full_name
                    lic.issued_to_email = email
                    lic.issued_at       = timezone.now()
                    lic.save()

                # ✅ Enhanced success message
                messages.success(
                    request,
                    f"✅ Church '{full_name}' registered and license key activated successfully! You can now log in."
                )
                return redirect('church:church_admin_login')

            except IntegrityError as e:
                form.add_error(None, f"A database error occurred: {e}")
            except Exception as e:
                form.add_error(None, f"An unexpected error occurred: {e}")

        else:
            messages.error(request, "Please correct the errors below.")

    else:
        form = ChurchRegistrationForm()

    return render(request, 'church/register_church.html', {'form': form})




class ChurchAdminLoginForm(forms.Form):
    email = forms.EmailField(label="Email")
    password = forms.CharField(widget=forms.PasswordInput, label="Password")

    def clean(self):
        cleaned_data = super().clean()
        email = cleaned_data.get("email")
        password = cleaned_data.get("password")

        if not email or not password:
            raise forms.ValidationError("Both email and password are required.")

        try:
            user = CustomUser.objects.get(email=email)
        except CustomUser.DoesNotExist:
            raise forms.ValidationError("User with this email does not exist.")

        if not user.is_church_admin:
            raise forms.ValidationError("This user is not a church admin.")

        if not user.check_password(password):
            raise forms.ValidationError("Invalid credentials.")

        cleaned_data['user'] = user
        return cleaned_data


church_admin_session_required = decorator_from_middleware(ChurchAdminSessionMiddleware)  # ✅ correct


@csrf_protect
def church_admin_login(request):
    """
    Logs a church‑admin in **unless their church is suspended**.
    Nothing else in the original flow has been changed.
    """
    if request.method == "POST":
        email    = request.POST.get("email")
        password = request.POST.get("password")
        user     = authenticate(request, email=email, password=password)

        if user is not None and user.is_church_admin:
            try:
                church = Church.objects.get(admin=user)

                # 🚫  NEW SUSPENSION GUARD
                if church.is_suspended:
                    messages.error(
                        request,
                        "This church account is suspended. Please contact support or "
                        "clear outstanding dues to regain access.",
                    )
                    return redirect("church:church_admin_login")

                # ✅  Normal login flow (unchanged)
                auth_login(request, user)

                request.session["church_admin_user_id"] = user.id
                request.session["church_id"] = church.id
                request.session.set_expiry(0)  # Session expires on browser close

                response = redirect("church:church_admin_dashboard")
                response.set_cookie(
                    "churchadmin_sessionid", request.session.session_key
                )

                return response

            except Church.DoesNotExist:
                messages.error(request, "Associated church not found.")
        else:
            messages.error(request, "Invalid email or password.")

    get_token(request)  # ensure CSRF token available for the form
    return render(request, "church/church_admin_login.html")




def church_admin_logout(request):
    request.session.flush()
    response = redirect('church:church_admin_login')
    response.delete_cookie('churchadmin_sessionid')
    messages.success(request, "Logged out successfully.")
    return response



church_admin_session_required = decorator_from_middleware(ChurchAdminSessionMiddleware)  # ✅ Correct


@login_required
def church_admin_dashboard(request):
    church_id = request.session.get("church_id")
    if not church_id:
        return redirect("church:church_admin_login")

    try:
        church = Church.objects.get(id=church_id)

        # ── Member stats ────────────────────────────────────────────────
        total_members = Member.objects.filter(church=church).count()

        category_counts = (
            Member.objects.filter(church=church)
            .values("category__name")
            .annotate(total=Count("id"))
        )
        category_stats = {"Men Society": 0, "Women Society": 0, "Youth Society": 0}
        for entry in category_counts:
            name = entry["category__name"]
            if name in category_stats:
                category_stats[name] = entry["total"]

        # Volunteers
        total_volunteers = Volunteer.objects.filter(church=church).count()

        # ── Upcoming events (next 5) ────────────────────────────────────
        upcoming_events = (
            Event.objects.filter(church=church, start_datetime__gte=timezone.now())
            .order_by("start_datetime")[:5]
        )

        # ── Upcoming live‑streams (next 5) ──────────────────────────────
        today = timezone.now().date()
        upcoming_streams = (
            LiveStream.objects.filter(
                church=church,
                is_active=True,
                date__gte=today,
            )
            .order_by("date")[:5]
        )

        # ── Chat rooms & announcements ─────────────────────────────────
        active_chat_rooms = ChatRoom.objects.filter(church=church)
        recent_announcements = (
            Announcement.objects.filter(church=church).order_by("-created_at")[:5]
        )

        # ── NEW: subscription due notice (within 7 days) ───────────────
        upcoming_due_date = None
        subscription = getattr(church, "subscription", None)
        if (
            subscription
            and subscription.is_active
            and subscription.next_due_date
            and today <= subscription.next_due_date <= today + timedelta(days=7)
        ):
            upcoming_due_date = subscription.next_due_date

        # ── Staff payments ──────────────────────────────────────────────
        current_month = today.month
        current_year = today.year

        # ✅ Total Paid This Month
        total_paid_this_month = (
            StaffPayment.objects.filter(
                staff__church=church,
                payment_month__year=current_year,
                payment_month__month=current_month,
            ).aggregate(total=Sum("amount"))["total"]
            or 0
        )

        context = {
            "church": church,
            "total_members": total_members,
            "category_stats": category_stats,
            "total_volunteers": total_volunteers,
            "upcoming_events": upcoming_events,
            "upcoming_streams": upcoming_streams,
            "active_chat_rooms": active_chat_rooms,
            "recent_announcements": recent_announcements,
            "upcoming_due_date": upcoming_due_date,
            "total_paid_this_month": total_paid_this_month,
        }

        return render(request, "church/church_admin_dashboard.html", context)

    except Church.DoesNotExist:
        raise Http404("Church not found.")




def church_admin_logout(request):
    logout(request)
    response = redirect('church:church_admin_login')  # or wherever your login view is
    response.delete_cookie('churchadmin_sessionid')   # 🧹 Clear the church admin session cookie
    return response



@login_required
def upload_church_image(request):
    # ✅ Step 1: Get church_id from session
    church_id = request.session.get('church_id')
    if not church_id:
        return redirect('church_admin_login')  # force re-login if session is lost

    # ✅ Step 2: Get the Church instance
    church = get_object_or_404(Church, id=church_id)

    # ✅ Step 3: Handle image upload
    if request.method == 'POST':
        form = ChurchImageForm(request.POST, request.FILES, instance=church)
        if form.is_valid():
            form.save()
            return redirect('church:church_admin_dashboard')
  
    else:
        form = ChurchImageForm(instance=church)

    return render(request, 'church/upload_image.html', {'form': form})


@login_required
def create_volunteer(request):
    if not request.user.is_church_admin:
        messages.error(request, "You are not authorized to create volunteers.")
        return redirect('church:church_admin_dashboard')

    # ✅ Safely get the admin’s church from session or request.user
    church_id = request.session.get('church_id')
    if not church_id:
        messages.error(request, "Your church session is missing. Please log in again.")
        return redirect('church:church_admin_login')

    try:
        church = Church.objects.get(id=church_id)
    except Church.DoesNotExist:
        messages.error(request, "Church not found.")
        return redirect('church:church_admin_login')

    if request.method == 'POST':
        form = VolunteerCreationForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.church = church  # ✅ Use valid church object
            user.is_volunteer = True
            user.save()

            Volunteer.objects.create(
                user=user,
                church=church,  # ✅ Use same valid church object
                can_manage_members=form.cleaned_data.get('can_manage_members', False),
                can_manage_streams=form.cleaned_data.get('can_manage_streams', False),
                can_manage_chats=form.cleaned_data.get('can_manage_chats', False),
                can_manage_announcements=form.cleaned_data.get('can_manage_announcements', False),
            )

            messages.success(request, "Volunteer created successfully!")
            return redirect('church:list_volunteers')
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = VolunteerCreationForm()

    return render(request, 'church/create_volunteer.html', {'form': form})



def volunteer_login(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')
        user = authenticate(request, email=email, password=password)

        if user is not None and hasattr(user, 'volunteer'):
            login(request, user)
            messages.success(request, 'Welcome, Volunteer!')
            return redirect('church:volunteer_dashboard')  # ✅ Ensure this exists in urls.py
        else:
            messages.error(request, 'Invalid credentials or not a volunteer.')
            return redirect('church:volunteer_login')  # 🔁 Stay on volunteer login if failed

    # ✅ Render login form on GET request
    return render(request, 'church/volunteer_login.html')




@login_required
def volunteer_dashboard(request):
    # Only volunteers allowed
    if not hasattr(request.user, 'volunteer'):
        raise Http404("Access denied.")

    church = request.user.volunteer.church

    # ── counts & data (unchanged) ────────────────────────────────────────
    total_members    = Member.objects.filter(church=church).count()
    total_volunteers = Volunteer.objects.filter(church=church).count()

    category_counts = (Member.objects
                       .filter(church=church)
                       .values('category__name')
                       .annotate(total=Count('id')))
    category_stats = {'Men Society': 0, 'Women Society': 0, 'Youth Society': 0}
    for row in category_counts:
        if row['category__name'] in category_stats:
            category_stats[row['category__name']] = row['total']

    upcoming_events  = (Event.objects
                        .filter(church=church, start_datetime__gte=timezone.now())
                        .order_by('start_datetime')[:5])

    upcoming_streams = (LiveStream.objects
                        .filter(church=church, date__gte=timezone.now())
                        .order_by('date')[:5])

    active_chat_rooms     = ChatRoom.objects.filter(church=church)
    recent_announcements  = (Announcement.objects
                             .filter(church=church)
                             .order_by('-created_at')[:5])

    # ── member chat (always fetched) ─────────────────────────────────────
    member_chats = (MemberChatMessage.objects
                    .filter(church=church, is_deleted=False)
                    .select_related("sender")
                    .order_by('-sent_at')[:30])

    context = {
        'church': church,
        'total_members': total_members,
        'category_stats': category_stats,
        'total_volunteers': total_volunteers,
        'upcoming_events': upcoming_events,
        'upcoming_streams': upcoming_streams,
        'active_chat_rooms': active_chat_rooms,
        'recent_announcements': recent_announcements,
        'member_chats': member_chats,
        'is_volunteer': hasattr(request.user, "volunteer"),
        # convenience flag for template
        'can_moderate': request.user.has_perm('church.moderate_message'),
    }
    return render(request, 'church/volunteer_dashboard.html', context)




def _get_user_church(user):
    if hasattr(user, 'churchadminprofile'):
        return user.churchadminprofile.church
    if hasattr(user, 'volunteer'):
        return user.volunteer.church
    raise PermissionDenied("User is not linked to a church.")



@login_required
def list_volunteers(request):
    church = _get_user_church(request.user)
    volunteers = Volunteer.objects.filter(church=church)
    return render(request, 'church/volunteers/list.html', {'volunteers': volunteers})



@login_required
def edit_volunteer(request, volunteer_id):
    church = _get_user_church(request.user)
    volunteer = get_object_or_404(Volunteer, id=volunteer_id, church=church)

    if request.method == 'POST':
        form = VolunteerForm(request.POST, instance=volunteer)
        if form.is_valid():
            form.save()
            return redirect('church:list_volunteers')
    else:
        form = VolunteerForm(instance=volunteer)

    return render(
        request,
        'church/edit_volunteer.html',  # ✅ MATCHES your actual file location
        {'form': form, 'volunteer': volunteer}
    )




@login_required
def delete_volunteer(request, volunteer_id):
    church = _get_user_church(request.user)
    volunteer = get_object_or_404(Volunteer, id=volunteer_id, church=church)

    if request.method == 'POST':
        volunteer.delete()
        return redirect('church:list_volunteers')

    return redirect('church:edit_volunteer', volunteer_id=volunteer.id)



class MemberRegistrationView(FormView):
    template_name = 'church/member_register.html'
    form_class = MemberRegistrationForm
    success_url = reverse_lazy('church:member_registration_success')  # Namespace-aware

    def form_valid(self, form):
        form.save()
        return super().form_valid(form)


class MemberRegistrationSuccessView(TemplateView):
    template_name = 'member_registration_success.html'



def home(request):
    return render(request, 'church/homepage.html') 


class MemberLoginView(View):
    def get(self, request):
        return render(request, 'church/member_login.html')

    def post(self, request):
        phone_number = request.POST.get('phone_number')
        password = request.POST.get('password')

        user = authenticate(request, username=phone_number, password=password)

        if user is not None and hasattr(user, 'member'):
            login(request, user)
            return redirect('church:member_dashboard')
        else:
            messages.error(request, 'Invalid phone number or password')
            return render(request, 'church/member_login.html')



@method_decorator(login_required, name="dispatch")
class MemberDashboardView(View):
    def get(self, request):
        now      = timezone.now()
        today    = now.date()
        tomorrow = today + timedelta(days=1)   # 24‑hour birthday window

        # 1. Guard – must have a Member profile
        if not hasattr(request.user, "member"):
            return render(
                request,
                "church/error.html",
                {"message": "You do not belong to any church profile."},
            )

        member  = request.user.member
        church  = member.church

        # 2. Upcoming events
        upcoming_events = (
            Event.objects.filter(is_public=True, start_datetime__gte=now, church=church)
                         .order_by("start_datetime")[:5]
        )

        # 3. Upcoming streams
        upcoming_streams = (
            LiveStream.objects.filter(is_active=True, date__gte=today, church=church)
                              .order_by("date")[:3]
        )

        # 4. Announcements
        announcements = (
            Announcement.objects.filter(church=church)
                                .filter(Q(is_public=True, user__isnull=True) | Q(user=request.user))
                                .order_by("-created_at")[:5]
        )

        # 5. Last 3 donations
        donations = (
            OnlineGiving.objects.filter(member=member, church=church)
                                .order_by("-created_at")[:3]
        )

        # 6. Today’s devotion
        reading_qs = BibleReadingPlan.objects.filter(church=church, reading_date=today).first()
        today_reading = None
        if reading_qs:
            try:
                parsed               = json.loads(reading_qs.content)
                reading_qs.parsed_verses = parsed.get("verses", [])
                reading_qs.commentary    = parsed.get("commentary", "")
                today_reading        = reading_qs
            except Exception:
                pass

        # 7. Prayer requests in last 24 h
        cutoff = now - timedelta(hours=24)
        prayer_requests = (
            PrayerRequest.objects.filter(member__member__church=church, created_at__gte=cutoff)
                                 .order_by("-created_at")[:5]
        )

        # 8. Stats
        member_categories = ["Men Society", "Women Society", "Youth Society"]
        category_counts_qs = (
            Member.objects.filter(church=church, category__name__in=member_categories)
                          .values("category__name")
                          .annotate(total=Count("id"))
        )
        category_stats = {label: 0 for label in member_categories}
        for entry in category_counts_qs:
            category_stats[entry["category__name"]] = entry["total"]

        total_volunteers = Volunteer.objects.filter(church=church).count()

        # 9. Upcoming birthdays (next 24 h)
        birthday_members = (
            Member.objects.filter(
                church=church,
                date_of_birth__isnull=False
            ).filter(
                Q(date_of_birth__month=today.month, date_of_birth__day=today.day) |
                Q(date_of_birth__month=tomorrow.month, date_of_birth__day=tomorrow.day)
            ).order_by("date_of_birth__month", "date_of_birth__day", "first_name")
        )

        # 🔟  Online members (NEW)
        online_members = (
            get_online_members()                   # all online users
            .filter(member__church=church)         # restrict to this church
            .select_related("member")              # avoid extra queries for picture/name
            .order_by("member__first_name")
        )

        return render(
            request,
            "church/member_dashboard.html",
            {
                "upcoming_events"   : upcoming_events,
                "upcoming_streams"  : upcoming_streams,
                "announcements"     : announcements,
                "donations"         : donations,
                "today_reading"     : today_reading,
                "prayer_requests"   : prayer_requests,
                "total_members"     : Member.objects.filter(church=church).count(),
                "category_stats"    : category_stats,
                "total_volunteers"  : total_volunteers,
                "member_categories" : member_categories,
                "birthday_members"  : birthday_members,
                "online_members"    : online_members,   # 🆕 passed to template
            },
        )




def member_logout(request):
    logout(request)
    return redirect('church:member_login')


class MemberProfileUpdateView(LoginRequiredMixin, UpdateView):
    model = Member
    fields = [
        'first_name',
        'last_name',
        'phone_number',
        'sex',
        'marital_status',
        'date_of_birth',
        'address',
        'profile_picture',  # ✅ Add this
    ]
    template_name = 'church/edit_member_profile.html'

    def get_object(self, queryset=None):
        return self.request.user.member

    def get_success_url(self):
        return reverse('church:member_dashboard')



# ✅ Mixin to restrict to church admin/volunteer
class IsChurchStaffMixin(UserPassesTestMixin):
    """
    Allows access to church admins and volunteers only.
    """

    def test_func(self):
        user = self.request.user
        return user.is_authenticated and (user.is_church_admin or hasattr(user, 'volunteer'))

    # 🔑  ADD THIS HELPER ↓↓↓
    def get_user_church(self):
        """
        Returns the Church instance associated with the current user
        (works for both ChurchAdminProfile and Volunteer).
        """
        user = self.request.user
        if hasattr(user, 'churchadminprofile'):
            return user.churchadminprofile.church
        elif hasattr(user, 'volunteer'):
            return user.volunteer.church
        # No church found → forbid access
        raise PermissionDenied("User does not belong to any church.")



class LiveStreamListView(LoginRequiredMixin, IsChurchStaffMixin, ListView):
    model = LiveStream
    template_name = 'church/livestream_list.html'
    context_object_name = 'upcoming_streams'

    def get_queryset(self):
        today = timezone.now().date()
        church = self.get_user_church()

        return LiveStream.objects.filter(
            church=church,
            is_active=True,          # Only show active ones
            date__gte=today          # All today and future
        ).order_by('date', 'start_time')  # Ordered by full schedule



# ✅ Admin/Staff - Create View
class LiveStreamCreateView(LoginRequiredMixin, IsChurchStaffMixin, CreateView):
    model = LiveStream
    template_name = 'church/add_livestream.html'
    fields = ['title', 'video_url', 'description', 'date', 'start_time', 'end_time', 'is_active']

    def form_valid(self, form):
        form.instance.church = self.get_user_church()
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('church:livestream_list')



# ✅ Admin/Staff - Update View
class LiveStreamUpdateView(LoginRequiredMixin, IsChurchStaffMixin, UpdateView):
    model = LiveStream
    fields = ['title', 'video_url', 'description', 'date', 'start_time', 'end_time', 'is_active']
    template_name = 'church/livestream_form.html'
    success_url = reverse_lazy('church:livestream_list')

    def get_queryset(self):
        return LiveStream.objects.filter(church=self.get_user_church())




# ✅ Admin/Staff - Delete View
class LiveStreamDeleteView(LoginRequiredMixin, IsChurchStaffMixin, DeleteView):
    model = LiveStream
    template_name = 'livestream/livestream_confirm_delete.html'
    success_url = reverse_lazy('church:livestream_list')

    def get_queryset(self):
        return LiveStream.objects.filter(church=self.get_user_church())



class MemberLiveStreamListView(LoginRequiredMixin, IsChurchMemberMixin, ListView):
    model = LiveStream
    template_name = 'church/member_livestream_list.html'
    context_object_name = 'livestreams'

    def get_queryset(self):
        now = timezone.now()
        return LiveStream.objects.filter(
            is_active=True,
            start_time__gte=now
        ).order_by('start_time')



# ✅ Member-facing: Watch livestream and join chat
class MemberLiveStreamDetailView(LoginRequiredMixin, DetailView):
    model = LiveStream
    template_name = 'church/member_livestream_detail.html'
    context_object_name = 'livestream'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        livestream = self.get_object()
        context['chatroom'] = getattr(livestream, 'chat_room', None)
        context['messages'] = context['chatroom'].messages.order_by('sent_at') if context['chatroom'] else []
        return context


# ✅ Member-facing: Post a chat message
@method_decorator(csrf_exempt, name='dispatch')
class PostMessageView(LoginRequiredMixin, View):
    def post(self, request, *args, **kwargs):
        chat_room_id = request.POST.get("chat_room_id")
        content = request.POST.get("content")
        chat_room = get_object_or_404(ChatRoom, id=chat_room_id)

        message = Message.objects.create(
            user=request.user,
            content=content,
            chat_room=chat_room,
            church=request.user.church
        )
        return JsonResponse({
            'username': request.user.username,
            'content': message.content
        })

    def get(self, request, *args, **kwargs):
        return JsonResponse({'error': 'GET not allowed'}, status=405)


class WatchLiveView(LoginRequiredMixin, TemplateView):
    template_name = 'church/watch_live.html'  # ← ✅ this is important

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        user = self.request.user

        if hasattr(user, 'member') and user.member:
            church = user.member.church
            stream = LiveStream.objects.filter(church=church, is_active=True).order_by('-date').first()
            context['stream'] = stream
            return context

        return redirect('church:church_admin_dashboard')



class LiveStreamDetailView(LoginRequiredMixin, DetailView):
    model = LiveStream
    template_name = 'livestream/livestream_detail.html'
    context_object_name = 'livestream'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        chatroom = ChatRoom.objects.filter(livestream=self.object).first()
        context['chatroom'] = chatroom
        context['messages'] = Message.objects.filter(chat_room=chatroom).order_by('sent_at') if chatroom else []
        return context



@method_decorator(login_required, name="dispatch")
class ChatRoomListView(ListView):
    """
    List all chat rooms that belong to the same church as the
    currently‑logged‑in user. Works for **admins + volunteers + members**.

    • Church‑admins          – user.churchadminprofile.church  
    • Volunteers             – user.volunteer.church  
    • Ordinary members       – user.member.church  
    """
    model = ChatRoom
    template_name = "church/chatroom_list.html"
    context_object_name = "chatrooms"

    def get_queryset(self):
        user = self.request.user

        if hasattr(user, "churchadminprofile"):
            self.church = user.churchadminprofile.church
        elif hasattr(user, "volunteer"):
            self.church = user.volunteer.church
        elif hasattr(user, "member"):
            self.church = user.member.church
        else:
            raise PermissionDenied("User is not linked to a church.")

        return ChatRoom.objects.filter(church=self.church)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user

        can_manage = False
        if hasattr(user, "churchadminprofile"):
            can_manage = True
        elif hasattr(user, "volunteer"):
            can_manage = getattr(user.volunteer, "can_manage_chats", False)

        context["can_manage"] = can_manage
        return context




def fetch_messages(request):
    chat_room_id = request.GET.get("chat_room_id")
    chat_room = ChatRoom.objects.filter(id=chat_room_id).first()
    messages = []

    if chat_room:
        for m in chat_room.messages.order_by("sent_at"):
            messages.append({
                "id": m.id,
                "username": m.user.username,
                "content": m.content,
                "timestamp": m.sent_at.strftime("%b %d, %Y %H:%M"),
                "role": "admin" if m.user.is_staff else "member",
                "can_moderate": request.user.is_staff
            })

    return JsonResponse({"messages": messages})



@csrf_exempt
@login_required
def edit_message(request):
    if request.method == "POST" and request.user.is_staff:
        message_id = request.POST.get("message_id")
        new_content = request.POST.get("new_content")
        message = Message.objects.filter(id=message_id).first()
        if message:
            message.content = new_content
            message.save()
    return JsonResponse({"status": "ok"})

@csrf_exempt
@login_required
def delete_message(request):
    if request.method == "POST" and request.user.is_staff:
        message_id = request.POST.get("message_id")
        Message.objects.filter(id=message_id).delete()
    return JsonResponse({"status": "ok"})



def _get_user_church(user):
    """
    Return the church this user belongs to (admin, volunteer, or member).
    Raise PermissionDenied if the user isn’t attached to any church.
    """
    if hasattr(user, "churchadminprofile"):
        return user.churchadminprofile.church
    if hasattr(user, "volunteer"):
        return user.volunteer.church
    if hasattr(user, "member"):
        return user.member.church
    raise PermissionDenied("User is not linked to a church.")


class CanManageChatMixin(UserPassesTestMixin):
    """
    Allow access when:
      • user is a church admin, OR
      • user is a volunteer *and* can_manage_chats flag is set.
    """

    def test_func(self):
        u = self.request.user
        if getattr(u, "is_church_admin", False):
            return True
        if getattr(u, "is_volunteer", False):
            # hasattr(u, 'volunteer') is safe because the flag is set
            return getattr(getattr(u, "volunteer", None), "can_manage_chats", False)
        return False

    # Optional: nicer 403 message
    def handle_no_permission(self):
        raise PermissionDenied("You don’t have permission to manage chat rooms.")


# ─── Chat‑room CRUD views ───────────────────────────────────────────────────
class ChatRoomCreateView(LoginRequiredMixin, CanManageChatMixin, CreateView):
    model         = ChatRoom
    form_class    = ChatRoomForm
    template_name = "church/chatroom_form.html"
    success_url   = reverse_lazy("church:chatroom_list")

    # pass “church” to the form (used to limit the livestream queryset, etc.)
    def get_form_kwargs(self):
        kwargs           = super().get_form_kwargs()
        kwargs["church"] = _get_user_church(self.request.user)
        return kwargs

    # set the ChatRoom’s church just before saving
    def form_valid(self, form):
        form.instance.church = _get_user_church(self.request.user)
        return super().form_valid(form)



class ChatRoomUpdateView(LoginRequiredMixin, UpdateView):
    model = ChatRoom
    form_class = ChatRoomForm
    template_name = 'church/chatroom_form.html'
    success_url = reverse_lazy('church:chatroom_list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['church'] = _get_user_church_generic(self.request.user)   # ✅
        return kwargs

    def form_valid(self, form):
        form.instance.church = _get_user_church_generic(self.request.user)  # ✅
        return super().form_valid(form)



class ChatRoomDetailView(LoginRequiredMixin, DetailView):
    model = ChatRoom
    template_name = "church/chatroom_detail.html"
    context_object_name = "chatroom"

    def get_queryset(self):
        user = self.request.user
        church = None
        if hasattr(user, "churchadminprofile"):
            church = user.churchadminprofile.church
        elif hasattr(user, "volunteer"):
            church = user.volunteer.church
        elif hasattr(user, "member"):
            church = user.member.church

        if not church:
            return ChatRoom.objects.none()

        return ChatRoom.objects.filter(church=church)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["messages"] = self.object.messages.order_by("sent_at")
        return ctx


class ChatRoomDeleteView(LoginRequiredMixin, CanManageChatMixin, DeleteView):
    model         = ChatRoom
    template_name = "church/chatroom_confirm_delete.html"
    success_url   = reverse_lazy("church:chatroom_list")

    def get_queryset(self):
        # can only delete rooms in your own church
        return ChatRoom.objects.filter(church=_get_user_church(self.request.user))




# ─── Helper mixin used by all three views ────────────────────
class EventPermissionMixin:
    """
    Resolves the church for either a Church‑Admin or a Volunteer.
    If the user has neither profile, access is denied.
    """

    def dispatch(self, request, *args, **kwargs):
        user = request.user
        self.church = None

        if hasattr(user, "churchadminprofile"):
            self.church = user.churchadminprofile.church
        elif hasattr(user, "volunteer"):
            self.church = user.volunteer.church

        if not self.church:
            return HttpResponseForbidden("You are not assigned to any church.")

        return super().dispatch(request, *args, **kwargs)

    # All list / update / delete queries are limited to the user’s church
    def get_queryset(self):
        qs = super().get_queryset()
        return qs.filter(church=self.church)




@method_decorator(login_required, name="dispatch")
class EventListView(EventPermissionMixin, ListView):
    """
    Lists events that belong to the logged‑in user’s church
    (works for either a Church‑Admin or a Volunteer).
    """
    model               = Event
    template_name       = "church/church_events.html"
    context_object_name = "events"
    paginate_by         = 20            # ← keep or remove pagination as you prefer

    def get_queryset(self):
        # EventPermissionMixin already set self.church
        return super().get_queryset().order_by("-start_datetime")




@method_decorator(login_required, name="dispatch")
class EventCreateView(EventPermissionMixin, CreateView):
    model         = Event
    form_class    = EventForm
    template_name = "church/event_form.html"
    success_url   = reverse_lazy("church:church_events")

    def form_valid(self, form):
        form.instance.church = self.church
        return super().form_valid(form)


@method_decorator(login_required, name="dispatch")
class EventUpdateView(EventPermissionMixin, UpdateView):
    model         = Event
    form_class    = EventForm
    template_name = "church/event_form.html"
    success_url   = reverse_lazy("church:church_events")


@method_decorator(login_required, name="dispatch")
class EventDeleteView(EventPermissionMixin, DeleteView):
    model         = Event
    template_name = "church/event_confirm_delete.html"
    success_url   = reverse_lazy("church:church_events")


@login_required
def notification_detail(request, pk):
    notification = get_object_or_404(Notification, pk=pk, user=request.user)

    # Mark as read automatically
    if not notification.is_read:
        notification.is_read = True
        notification.save()

    return render(request, 'church/notification_detail.html', {'notification': notification})
 

@login_required
def all_notifications(request):
    user = request.user

    try:
        if hasattr(user, 'churchadminprofile'):
            church = user.churchadminprofile.church
        elif hasattr(user, 'member'):
            church = user.member.church
        else:
            return render(request, 'church/error.html', {
                'message': "You do not belong to any church profile."
            })
    except Exception:
        return render(request, 'church/error.html', {
            'message': "Profile access error."
        })

    notifications = Notification.objects.filter(user=user).order_by('-created_at')

    return render(request, 'church/all_notifications.html', {
        'notifications': notifications
    })


@require_POST
@login_required
def mark_all_notifications_as_read(request):
    user = request.user
    try:
        if hasattr(user, 'churchadminprofile'):
            church = user.churchadminprofile.church
        elif hasattr(user, 'member'):
            church = user.member.church
        else:
            return HttpResponseForbidden("You do not belong to any church profile.")
    except Exception:
        return HttpResponseForbidden("Profile access error.")

    unread_qs = Notification.objects.filter(user=user, is_read=False)
    count = unread_qs.count()
    if count > 0:
        unread_qs.update(is_read=True)
        messages.success(request, f"{count} notifications marked as read.")
    else:
        messages.info(request, "All notifications are already marked as read.")

    return redirect('church:all_notifications')



@login_required
def member_notifications(request):
    user = request.user

    try:
        church = user.member.church
    except AttributeError:
        return render(request, 'church/error.html', {
            'message': "You do not belong to any church profile."
        })

    notifications = Announcement.objects.filter(
        Q(church=church),
        Q(is_public=True, user__isnull=True) | Q(user=user)
    ).order_by('-created_at')

    return render(request, 'church/member_notification_list.html', {
        'notifications': notifications
    })




class NotificationListView(LoginRequiredMixin, ListView):
    model = Notification
    template_name = 'church/all_notifications.html'
    context_object_name = 'notifications'

    def get_queryset(self):
        u = self.request.user
        if hasattr(u, 'churchadminprofile'):
            return Notification.objects.filter(
                church=u.churchadminprofile.church
            ).order_by('-created_at')
        elif hasattr(u, 'member'):          # ✅ was churchmemberprofile
            return Notification.objects.filter(
                church=u.member.church
            ).order_by('-created_at')
        return Notification.objects.none()



@login_required
def check_new_notifications(request):
    unread_notifications = Notification.objects.filter(user=request.user, is_read=False).order_by('-created_at')
    notifications = list(unread_notifications.values('id', 'message'))
    return JsonResponse({
        'new_notifications': unread_notifications.exists(),
        'notifications': notifications,
        'unread_count': unread_notifications.count(),
    })



@login_required
def members(request):
    church = _get_user_church_generic(request.user)
    members = Member.objects.filter(church=church)
    return render(request, 'church/members.html', {'members': members})



# Re‑usable helper to get the church
def _get_user_church(user):
    """
    Return the church this user belongs to.
    Works for church admins and volunteers.
    Raises PermissionDenied if the user has no church link.
    """
    if hasattr(user, 'churchadminprofile'):
        return user.churchadminprofile.church
    if hasattr(user, 'volunteer'):
        return user.volunteer.church
    raise PermissionDenied("User is not linked to a church.")



# List all announcements and personal notifications created by church admin
class AdminAnnouncementListView(LoginRequiredMixin, ListView):
    model = Announcement
    template_name = 'church/church_announcement_list.html'
    context_object_name = 'announcements'

    def get_queryset(self):
        church = _get_user_church(self.request.user)
        return Announcement.objects.filter(church=church).order_by("-created_at")


class AnnouncementCreateView(LoginRequiredMixin, CreateView):
    model = Announcement
    fields = ['title', 'content', 'event', 'user', 'is_public']
    template_name = 'church/announcement_form.html'
    success_url = reverse_lazy('church:church_announcement_list')

    def form_valid(self, form):
        form.instance.church = _get_user_church(self.request.user)
        form.instance.created_by = self.request.user
        return super().form_valid(form)


class AnnouncementUpdateView(LoginRequiredMixin, UpdateView):
    model = Announcement
    fields = ['title', 'content', 'event', 'user', 'is_public']
    template_name = 'church/announcement_form.html'
    success_url = reverse_lazy('church:church_announcement_list')

    def get_queryset(self):
        return Announcement.objects.filter(church=_get_user_church(self.request.user))


class AnnouncementDeleteView(LoginRequiredMixin, DeleteView):
    model = Announcement
    template_name = 'church/announcement_confirm_delete.html'
    success_url = reverse_lazy('church:church_announcement_list')

    def get_queryset(self):
        return Announcement.objects.filter(church=_get_user_church(self.request.user))



class MemberAnnouncementListView(LoginRequiredMixin, ListView):
    model = Announcement
    template_name = 'church/member_announcement_list.html'
    context_object_name = 'announcements'

    def get_queryset(self):
        user = self.request.user
        church = user.member.church
        return Announcement.objects.filter(
            Q(is_public=True, user__isnull=True, church=church) |  # Public announcements from admin
            Q(user=user, is_public=False, church=church)           # Personal/private ones
        ).order_by('-created_at').distinct()



class MemberAnnouncementDetailView(LoginRequiredMixin, DetailView):
    model = Announcement
    template_name = 'church/member_announcement_detail.html'
    context_object_name = 'announcement'

    def get_queryset(self):
        user = self.request.user
        if hasattr(user, 'member'):
            church = user.member.church
            return Announcement.objects.filter(
                church=church
            ).filter(
                Q(is_public=True, user__isnull=True) | Q(user=user)
            )
        else:
            return Announcement.objects.none()




class MarkAnnouncementReadView(LoginRequiredMixin, View):
    def get(self, request, pk):
        ann = get_object_or_404(Announcement, pk=pk, user=request.user, is_public=False)
        ann.is_read = True
        ann.save()
        return redirect('member_announcements')



@method_decorator(church_admin_session_required, name='dispatch')
class StewardshipRecordListView(ListView):
    model = StewardshipRecord
    template_name = 'church/stewardship_list.html'
    context_object_name = 'page_obj'

    def get_church(self):
        church_id = self.request.session.get('church_id')
        if not church_id:
            raise Http404("Church not found in session.")
        return Church.objects.get(id=church_id)

    def get_queryset(self):
        church = self.get_church()
        qs = StewardshipRecord.objects.filter(
            church=church, is_trashed=False
        ).order_by('-date', '-created_at')

        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')
        trans_type = self.request.GET.get('transaction_type')
        category = self.request.GET.get('category')

        if start_date:
            qs = qs.filter(date__gte=start_date)
        if end_date:
            qs = qs.filter(date__lte=end_date)
        if trans_type:
            qs = qs.filter(transaction_type=trans_type)
        if category:
            qs = qs.filter(category=category)

        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        church = self.get_church()
        all_records = self.get_queryset()

        paginator = Paginator(all_records, 10)
        page = self.request.GET.get('page')
        try:
            paginated_records = paginator.page(page)
        except PageNotAnInteger:
            paginated_records = paginator.page(1)
        except EmptyPage:
            paginated_records = paginator.page(paginator.num_pages)

        for record in paginated_records:
            record.logs = AuditLog.objects.filter(stewardship_record=record).order_by('-timestamp')

        opening_balance = StewardshipRecord.objects.filter(
            church=church, category='opening_balance', is_trashed=False
        ).aggregate(total=Sum('amount'))['total'] or 0

        income = all_records.filter(transaction_type='IN').exclude(category='opening_balance').aggregate(total=Sum('amount'))['total'] or 0
        expenditure = all_records.filter(transaction_type='EX').aggregate(total=Sum('amount'))['total'] or 0
        current_balance = opening_balance + income - expenditure

        monthly_summary = defaultdict(lambda: {'total_income': 0, 'total_expenditure': 0})
        running_balance = []
        labels = []
        income_data = []
        expenditure_data = []
        balance = opening_balance

        for record in all_records.order_by('date'):
            month_key = record.date.strftime('%Y-%m')

            if record.transaction_type == 'IN':
               monthly_summary[month_key]['total_income'] += float(record.amount)

            elif record.transaction_type == 'EX':
                monthly_summary[month_key]['total_expenditure'] +=    float(record.amount)


        sorted_summary = OrderedDict(sorted(monthly_summary.items()))
        formatted_summary = []
        for month_key, data in sorted_summary.items():
            date_obj = datetime.strptime(month_key, '%Y-%m')
            label = date_obj.strftime('%b %Y')
            labels.append(label)
            income_data.append(data['total_income'])
            expenditure_data.append(data['total_expenditure'])
            formatted_summary.append({
                'month': label,
                'income': data['total_income'],
                'expense': data['total_expenditure'],
            })

        category_totals = all_records.values('category').annotate(total=Sum('amount'))

        querydict = self.request.GET.copy()
        querydict.pop('page', None)
        querystring = querydict.urlencode()

        context.update({
            'page_obj': paginated_records,
            'opening_balance': opening_balance,
            'total_income': income,
            'total_expenditure': expenditure,
            'current_balance': current_balance,
            'monthly_summary': formatted_summary,
            'chart_labels': labels,
            'chart_income': income_data,
            'chart_expenditure': expenditure_data,
            'running_balance': running_balance,
            'category_totals': category_totals,
            'trash_records': StewardshipRecord.objects.filter(church=church, is_trashed=True).order_by('-date', '-created_at'),
            'querystring': querystring,
        })
        return context



@login_required
def add_stewardship_record(request):
    church_id = request.session.get('church_id')
    if not church_id:
        return redirect('church:church_admin_login')

    church = get_object_or_404(Church, id=church_id)

    if request.method == 'POST':
        print("Form submitted")  # Debugging: Confirm POST is received
        form = StewardshipForm(request.POST)
        if form.is_valid():
            record = form.save(commit=False)
            record.church = church
            record.created_at = timezone.now()
            record.is_trashed = False
            record.save()

            # ✅ Create AuditLog entry
            AuditLog.objects.create(
                stewardship_record=record,
                user=request.user,
                action='Created',
                message=f"Record created: ₹{record.amount} for {record.category} on {record.date}"
)

            messages.success(request, "Record added successfully!")
            form = StewardshipForm()  # Reset the form
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = StewardshipForm()

    return render(request, 'church/add_record.html', {'form': form})


@login_required
def generate_stewardship_pdf(request):
    user = request.user
    church = None

    church_id = request.session.get('church_id')
    if church_id:
        church = get_object_or_404(Church, id=church_id)
    elif user.is_superuser:
        church_id = request.GET.get('church_id')
        if not church_id:
            return HttpResponse("Superuser must provide ?church_id=", status=400)
        church = get_object_or_404(Church, id=church_id)

    if not church:
        return HttpResponseNotFound("Church not found.")

    records = StewardshipRecord.objects.filter(church=church, is_trashed=False).order_by('date')

    opening_balance = StewardshipRecord.objects.filter(
        church=church, category='opening_balance', is_trashed=False
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

    income = StewardshipRecord.objects.filter(
        church=church, transaction_type='IN', is_trashed=False
    ).exclude(category='opening_balance').aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

    expenditure = StewardshipRecord.objects.filter(
        church=church, transaction_type='EX', is_trashed=False
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

    current_balance = opening_balance + income - expenditure

    image_url = church.profile_picture.url if church.profile_picture else None

    context = {
        'records': records,
        'church_name': church.church_name,
        'church_image_url': image_url,
        'opening_balance': opening_balance,
        'total_income': income,
        'total_expenditure': expenditure,
        'current_balance': current_balance,
        'now': timezone.now(),
    }

    html_string = get_template('church/stewardship_pdf.html').render(context)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'filename="stewardship_report.pdf"'

    fd, path = tempfile.mkstemp(suffix=".pdf")
    try:
        HTML(string=html_string, base_url=request.build_absolute_uri('/')).write_pdf(path)
        with open(path, 'rb') as pdf_file:
            response.write(pdf_file.read())
    finally:
        os.close(fd)
        os.remove(path)

    return response


@login_required
def export_stewardship_to_excel(request):
    church = request.user.churchadminprofile.church
    records = StewardshipRecord.objects.filter(church=church, is_trashed=False).order_by('date')

    wb = Workbook()
    ws = wb.active
    ws.title = "Stewardship Records"

    church_name = str(church.name).upper() if hasattr(church, 'name') else "CHURCH STEWARDSHIP REPORT"
    ws.merge_cells('A1:E1')
    title_cell = ws['A1']
    title_cell.value = f"{church_name} - STEWARDSHIP REPORT"
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center')

    headers = ['Date', 'Type', 'Category', 'Amount', 'Description']
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center')

    ws.append(headers)
    for col_num, column_title in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    opening_balance = StewardshipRecord.objects.filter(
        church=church, category='opening_balance', is_trashed=False
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

    income_total = Decimal('0.00')
    expenditure_total = Decimal('0.00')

    for idx, record in enumerate(records, start=3):
        amount_decimal = Decimal(record.amount)
        amount_str = f"₹{float(amount_decimal):,.2f}"

        row = [
            record.date.strftime('%Y-%m-%d'),
            record.get_transaction_type_display(),
            record.get_category_display(),
            amount_str,
            record.description or '-'
        ]
        ws.append(row)

        fill_color = "C6EFCE" if record.transaction_type == 'IN' else "F4CCCC"
        for col_num in range(1, 6):
            cell = ws.cell(row=idx, column=col_num)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        if record.transaction_type == 'IN' and record.category != 'opening_balance':
            income_total += amount_decimal
        elif record.transaction_type == 'EX':
            expenditure_total += amount_decimal

    current_balance = opening_balance + income_total - expenditure_total

    ws.append([])

    summary_start_row = ws.max_row + 1
    summary_data = [
        ('Opening Balance', opening_balance, 'FFF2CC'),
        ('Total Income', income_total, 'C6EFCE'),
        ('Total Expenditure', expenditure_total, 'F4CCCC'),
        ('Current Balance', current_balance, 'BDD7EE'),
    ]

    for idx, (label, amount, color) in enumerate(summary_data, start=summary_start_row):
        ws.cell(row=idx, column=4).value = label
        ws.cell(row=idx, column=5).value = f"₹{float(amount):,.2f}"
        ws.cell(row=idx, column=4).font = Font(bold=True)
        ws.cell(row=idx, column=5).font = Font(bold=True)
        ws.cell(row=idx, column=4).alignment = Alignment(horizontal='center')
        ws.cell(row=idx, column=5).alignment = Alignment(horizontal='center')
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws.cell(row=idx, column=4).fill = fill
        ws.cell(row=idx, column=5).fill = fill

    for col_idx, col_cells in enumerate(ws.columns, 1):
        max_length = 0
        for cell in col_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max_length + 2

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=stewardship_records.xlsx'
    wb.save(response)
    return response



@login_required
def generate_filtered_pdf(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    transaction_type = request.GET.get('transaction_type')
    category = request.GET.get('category')

    church = request.user.churchadminprofile.church

    records = StewardshipRecord.objects.filter(church=church, is_trashed=False).order_by('date')

    if start_date:
        records = records.filter(date__gte=start_date)
    if end_date:
        records = records.filter(date__lte=end_date)
    if transaction_type:
        records = records.filter(transaction_type=transaction_type)
    if category:
        records = records.filter(category=category)

    income = records.filter(transaction_type='IN').exclude(category='opening') \
        .aggregate(total=Sum('amount'))['total'] or 0

    expenditure = records.filter(transaction_type='EX') \
        .aggregate(total=Sum('amount'))['total'] or 0

    current_balance = income - expenditure

    image_url = church.profile_picture.url if church.profile_picture else None

    context = {
        'records': records,
        'church_name': church.church_name,
        'church_image_url': image_url,
        'total_income': income,
        'total_expenditure': expenditure,
        'current_balance': current_balance,
        'now': timezone.now(),
        'generated_at': timezone.now(),
        'filtered': True,
        'filters': {
            'start_date': start_date,
            'end_date': end_date,
            'transaction_type': transaction_type,
            'category': category,
        }
    }

    html_string = get_template('church/filtered_report_pdf.html').render(context)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="filtered_stewardship_report.pdf"'

    fd, path = tempfile.mkstemp(suffix=".pdf")
    try:
        HTML(string=html_string, base_url=request.build_absolute_uri('/')).write_pdf(path)
        with open(path, 'rb') as pdf_file:
            response.write(pdf_file.read())
    finally:
        os.close(fd)
        os.remove(path)

    return response



@login_required
def export_filtered_excel(request):
    church = request.user.churchadminprofile.church
    church_name = church.name.upper() if hasattr(church, 'name') else "CHURCH STEWARDSHIP REPORT"

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    transaction_type = request.GET.get('transaction_type')
    category = request.GET.get('category')

    records = StewardshipRecord.objects.filter(church=church, is_trashed=False)

    if start_date:
        records = records.filter(date__gte=start_date)
    if end_date:
        records = records.filter(date__lte=end_date)
    if transaction_type:
        records = records.filter(transaction_type=transaction_type)
    if category:
        records = records.filter(category=category)

    total_income = sum(r.amount for r in records if r.transaction_type == 'IN' and r.category != 'opening')
    total_expenditure = sum(r.amount for r in records if r.transaction_type == 'EX')
    current_balance = total_income - total_expenditure

    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet("Filtered Stewardship Records")

    worksheet.set_column(0, 4, 20)

    worksheet.merge_range('A1:E1', f"{church_name} - STEWARDSHIP REPORT", workbook.add_format({
        'bold': True, 'font_size': 14, 'align': 'center', 'valign': 'vcenter',
        'bg_color': '#D9E1F2'
    }))

    header_format = workbook.add_format({
        'bold': True, 'bg_color': '#D3D3D3', 'border': 1,
        'align': 'center', 'valign': 'vcenter'
    })

    income_date_format = workbook.add_format({
        'num_format': 'yyyy-mm-dd', 'border': 1, 'bg_color': '#DFF0D8',
        'align': 'center', 'valign': 'vcenter'
    })
    income_text_format = workbook.add_format({
        'border': 1, 'bg_color': '#DFF0D8',
        'align': 'center', 'valign': 'vcenter'
    })
    income_currency_format = workbook.add_format({
        'num_format': '₹#,##0.00', 'border': 1, 'bg_color': '#DFF0D8',
        'align': 'center', 'valign': 'vcenter'
    })

    expenditure_date_format = workbook.add_format({
        'num_format': 'yyyy-mm-dd', 'border': 1, 'bg_color': '#F2DEDE',
        'align': 'center', 'valign': 'vcenter'
    })
    expenditure_text_format = workbook.add_format({
        'border': 1, 'bg_color': '#F2DEDE',
        'align': 'center', 'valign': 'vcenter'
    })
    expenditure_currency_format = workbook.add_format({
        'num_format': '₹#,##0.00', 'border': 1, 'bg_color': '#F2DEDE',
        'align': 'center', 'valign': 'vcenter'
    })

    total_label_format = workbook.add_format({
        'bold': True, 'align': 'right', 'valign': 'vcenter'
    })
    total_value_format = workbook.add_format({
        'bold': True, 'num_format': '₹#,##0.00', 'align': 'center', 'valign': 'vcenter'
    })

    headers = ["Date", "Type", "Category", "Amount", "Description"]
    for col_num, header in enumerate(headers):
        worksheet.write(1, col_num, header, header_format)

    for row_num, record in enumerate(records, start=2):
        is_income = record.transaction_type == 'IN'
        date_format = income_date_format if is_income else expenditure_date_format
        text_format = income_text_format if is_income else expenditure_text_format
        currency_format = income_currency_format if is_income else expenditure_currency_format

        worksheet.write_datetime(row_num, 0, record.date, date_format)
        worksheet.write(row_num, 1, record.get_transaction_type_display(), text_format)
        worksheet.write(row_num, 2, record.get_category_display(), text_format)
        worksheet.write_number(row_num, 3, float(record.amount), currency_format)
        worksheet.write(row_num, 4, record.description or "-", text_format)

    summary_start_row = len(records) + 3

    worksheet.write(summary_start_row, 2, "Total Income:", total_label_format)
    worksheet.write_number(summary_start_row, 3, float(total_income), total_value_format)

    worksheet.write(summary_start_row + 1, 2, "Total Expenditure:", total_label_format)
    worksheet.write_number(summary_start_row + 1, 3, float(total_expenditure), total_value_format)

    worksheet.write(summary_start_row + 2, 2, "Current Balance:", total_label_format)
    worksheet.write_number(summary_start_row + 2, 3, float(current_balance), total_value_format)

    workbook.close()
    output.seek(0)

    filename = 'filtered_stewardship_records.xlsx'
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    return response



class StewardshipRecordUpdateView(UpdateView):
    model = StewardshipRecord
    form_class = StewardshipForm
    template_name = 'church/edit_record.html'
    success_url = reverse_lazy('church:stewardship_list')

    def form_valid(self, form):
        response = super().form_valid(form)

        # Create an audit log entry
        AuditLog.objects.create(
            stewardship_record=self.object,
            user=self.request.user,
            action='Updated',
            message='Record updated successfully.'
        )

        return response


# ✅ Soft Delete (Function-Based View)
@login_required
def soft_delete_stewardship_record(request, pk):
    church = request.user.churchadminprofile.church
    record = get_object_or_404(StewardshipRecord, pk=pk, church=church)
    record.is_trashed = True
    record.trashed_at = timezone.now()
    record.save()

    AuditLog.objects.create(
        stewardship_record=record,
        user=request.user,
        action='Soft Delete',
        message='Marked as trashed.'
    )

    messages.success(request, "Record moved to trash.")
    return redirect('church:stewardship_list')


# ✅ Trash List
def stewardship_trash_list(request):
    church = request.user.churchadminprofile.church
    trashed_records = StewardshipRecord.objects.filter(
        church=church,
        is_trashed=True
    ).order_by('-trashed_at')

    print("🚮 Trashed record IDs:", list(trashed_records.values_list('id', flat=True)))

    return render(request, 'church/stewardship_trash_list.html', {
        'trashed_records': trashed_records
    })


# ✅ Restore
@require_POST
@login_required
def restore_stewardship_record(request, pk):
    church = request.user.churchadminprofile.church
    record = get_object_or_404(StewardshipRecord, pk=pk, is_trashed=True, church=church)
    record.is_trashed = False
    record.trashed_at = None
    record.save()

    AuditLog.objects.create(
        stewardship_record=record,
        user=request.user,
        action='Restore',
        message='Restored from trash.'
    )

    messages.success(request, "Record successfully restored.")
    return redirect('church:stewardship_trash_list')


# ✅ Empty Trash
@require_POST
@login_required
def empty_stewardship_trash(request):
    church = request.user.churchadminprofile.church
    records = StewardshipRecord.objects.filter(church=church, is_trashed=True)
    count = records.count()

    for record in records:
        AuditLog.objects.create(
            stewardship_record=record,
            user=request.user,
            action='Empty Trash',
            message='Permanently deleted from trash.'
        )
        record.hard_delete()  # ✅ Proper permanent delete

    messages.success(request, f"Emptied {count} trashed records.")
    return redirect('church:stewardship_trash_list')


# ✅ Class-Based Soft Delete via DeleteView
class StewardshipRecordDeleteView(DeleteView):
    model = StewardshipRecord
    context_object_name = 'record'
    template_name = 'church/stewardship/confirm_soft_delete.html'
    success_url = reverse_lazy('church:stewardship_list')

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        self.object.is_trashed = True
        self.object.trashed_at = timezone.now()
        self.object.save()

        AuditLog.objects.create(
            stewardship_record=self.object,
            user=request.user,
            action='Soft Delete (CBV)',
            message='Moved to trash via DeleteView'
        )

        messages.success(request, "Record moved to trash.")
        return redirect(self.success_url)


@login_required(login_url='church:church_admin_login')
def audit_log_view(request):
    church = request.user.churchadminprofile.church
    logs = AuditLog.objects.filter(
        stewardship_record__church=church
    ).select_related('user', 'stewardship_record').order_by('-timestamp')

    return render(request, 'church/audit_log.html', {
        'logs': logs
    })



@login_required
def stewardship_list(request):
    church = request.user.churchadminprofile.church
    records = StewardshipRecord.objects.filter(church=church, is_trashed=False).order_by('-date')

    # Prefetch audit logs per record (efficient)
    records = records.prefetch_related(
        Prefetch('auditlog_set', queryset=AuditLog.objects.order_by('-timestamp'), to_attr='logs')
    )

    return render(request, 'church/stewardship_list.html', {
        'records': records,
    })




@login_required
def create_sunday_income_receipt(request):
    user = request.user

    if not user.is_church_admin:
        return redirect('church_admin_login')

    try:
        admin_profile = ChurchAdminProfile.objects.get(user=user)
        church = admin_profile.church
    except ChurchAdminProfile.DoesNotExist:
        return redirect('church_admin_login')

    initial_data = {}

    if request.method == 'GET':
        auto_fill = request.GET.get('auto_fill') == '1'
        sunday_date = request.GET.get('sunday_date')

        if auto_fill:
            last_sunday = now().date() - timedelta(days=now().weekday() + 1)
            receipt = SundayIncomeReceipt.objects.filter(
                church=church,
                created_at__date=last_sunday
            ).order_by('-created_at').first()
        elif sunday_date:
            receipt = SundayIncomeReceipt.objects.filter(
                church=church,
                created_at__date=sunday_date
            ).order_by('-created_at').first()
        else:
            receipt = None

        if receipt:
            initial_data = {
                'name': receipt.name,
                'category': receipt.category,
                'amount': receipt.amount,
                'receiver_name': receipt.receiver_name or church.church_name,
                'thank_you_message': receipt.thank_you_message or "Thank you for your generosity!",
            }
        else:
            initial_data = {
                'receiver_name': church.church_name,
                'thank_you_message': "Thank you for your generosity",
            }

        form = SundayIncomeReceiptForm(initial=initial_data)

    elif request.method == 'POST':
        form = SundayIncomeReceiptForm(request.POST, request.FILES)
        if form.is_valid():
            receipt = form.save(commit=False)
            receipt.church = church
            if not receipt.receiver_name:
                receipt.receiver_name = church.church_name
            if not receipt.thank_you_message:
                receipt.thank_you_message = "Thank you for your generosity"
            receipt.created_at = now()
            receipt.save()

            # ✅ Send SMS if phone_number exists
            if receipt.phone_number:
                send_sms(
                    receipt.phone_number,
                    f"Thank you for your {receipt.get_category_display()} of ₹{receipt.amount} "
                    f"to {receipt.church.church_name} on {now().strftime('%d %b %Y %I:%M %p')}."
                )

            return redirect('church:print_sunday_income_receipt', receipt_id=receipt.id)

    context = {
        'form': form,
        'church_name': church.church_name,
        'church_logo': church.profile_picture.url if church.profile_picture else None,
        'current_datetime': now(),
    }
    return render(request, 'church/create_sunday_income_receipt.html', context)



@login_required
def print_sunday_income_receipt(request, receipt_id):
    user = request.user

    # Ensure only church admins can access
    if not user.is_church_admin:
        return redirect('church_admin_login')

    try:
        receipt = SundayIncomeReceipt.objects.get(id=receipt_id, church=user.churchadminprofile.church)
        church = user.churchadminprofile.church
    except SundayIncomeReceipt.DoesNotExist:
        return HttpResponse("Receipt not found.", status=404)

    context = {
        'receipt': receipt,
        'church_name': church.church_name,
        'church_logo': church.profile_picture.url if church.profile_picture else None,
    }

    return render(request, 'church/income_receipt_print.html', context)


class PrintOnlineGivingReceiptView(View):
    def get(self, request, receipt_id):
        receipt = get_object_or_404(OnlineGiving, id=receipt_id)

        if hasattr(request.user, 'churchadminprofile'):
            church = request.user.churchadminprofile.church
        elif hasattr(request.user, 'member'):            # ✅ was memberprofile
            church = request.user.member.church
            if receipt.member != request.user.member:
                return HttpResponseForbidden("This receipt does not belong to you.")
        else:
            return HttpResponseForbidden("Access denied.")

        if receipt.church != church:
            return HttpResponseForbidden("Access denied.")

        return render(request, 'church/online_giving_receipt.html', {
            'receipt': receipt,
            'church': church,
            'is_print_view': True,
        })



@method_decorator(login_required, name='dispatch')
class OnlineGivingView(View):
    def get(self, request):
        form = OnlineGivingForm()

        # ✅ Identify the correct church
        church = None
        if hasattr(request.user, 'member'):
            church = request.user.member.church
        elif hasattr(request.user, 'churchadminprofile'):
            church = request.user.churchadminprofile.church

        return render(request, 'church/online_giving_form.html', {
            'form': form,
            'church': church,
        })

    def post(self, request):
        form = OnlineGivingForm(request.POST)
        if form.is_valid():
            giving = form.save(commit=False)
            if hasattr(request.user, 'member'):
                giving.church = request.user.member.church
                giving.member = request.user.member
            elif hasattr(request.user, 'churchadminprofile'):
                giving.church = request.user.churchadminprofile.church
            giving.save()

            # ✅ Send SMS
            send_sms(
                giving.phone_number,
                f"🙏 Thank you for your {giving.get_category_display()} of ₹{giving.amount} "
                f"to {giving.church.church_name} on {localtime(giving.created_at).strftime('%d %b %Y %I:%M %p')}. "
                f"Receipt No: {giving.receipt_number}"
            )

            return redirect('church:view_online_giving_receipt', receipt_id=giving.id)

        # Retry form with context
        church = None
        if hasattr(request.user, 'member'):
            church = request.user.member.church
        elif hasattr(request.user, 'churchadminprofile'):
            church = request.user.churchadminprofile.church

        return render(request, 'church/online_giving_form.html', {
            'form': form,
            'church': church,
        })





class OnlineGivingCreateView(CreateView):
    model = OnlineGiving
    form_class = OnlineGivingForm
    template_name = 'church/online_giving_form.html'

    def form_valid(self, form):
        giving = form.save(commit=False)

        if hasattr(self.request.user, "member"):
            giving.church = self.request.user.member.church
            giving.member = self.request.user.member
        elif hasattr(self.request.user, "churchadminprofile"):
            giving.church = self.request.user.churchadminprofile.church
        else:
            return HttpResponseForbidden("Access denied")

        giving.save()

        send_sms(
            giving.phone_number,
            f"🙏 Thank you for your {giving.get_category_display()} of ₹{giving.amount} "
            f"to {giving.church.church_name} on "
            f"{localtime(giving.created_at).strftime('%d %b %Y %I:%M %p')}. "
            f"Receipt No: {giving.receipt_number}"
        )

        return redirect('church:view_online_giving_receipt', receipt_id=giving.pk)




@method_decorator(login_required, name='dispatch')
class OnlineGivingReceiptView(View):
    def get(self, request, receipt_id):
        receipt = get_object_or_404(OnlineGiving, id=receipt_id)

        # Determine which user is accessing
        user = request.user

        if hasattr(user, 'member'):
            if receipt.member != user.member:
                return HttpResponseForbidden("This receipt does not belong to you.")
            church = user.member.church
        elif hasattr(user, 'churchadminprofile'):
            church = user.churchadminprofile.church
        else:
            return HttpResponseForbidden("Access denied.")

        # Check receipt church
        if receipt.church != church:
            return HttpResponseForbidden("Access denied.")

        context = {
            'receipt': receipt,
            'church': church,
            'is_print_view': 'pdf' in request.path,  # True if printing, False if just viewing
            'dashboard_url': reverse('church:member_dashboard') 
                if hasattr(user, 'member') 
                else reverse('church:church_admin_dashboard'),
        }

        return render(request, 'church/online_giving_receipt.html', context)




@login_required
def sunday_income_receipt_list(request):
    user = request.user
    church = user.churchadminprofile.church

    # Fetch all SundayIncomeReceipt and OnlineGiving records for this church
    sunday_receipts = SundayIncomeReceipt.objects.filter(church=church, is_deleted=False)
    online_givings = OnlineGiving.objects.filter(church=church)

    # Annotate sunday_receipts with source_type
    for receipt in sunday_receipts:
        receipt.source_type = 'sunday'

    # Annotate online_givings with expected fields for display
    for giving in online_givings:
        giving.name = giving.giver_name
        giving.receiver_name = church.upi_id or church.church_name
        giving.category = giving.category or 'unknown'
        giving.thank_you_message = giving.thank_you_message
        giving.source_type = 'online'

    # Combine and sort by date (descending)
    all_receipts = sorted(
        chain(sunday_receipts, online_givings),
        key=lambda r: localtime(r.created_at),
        reverse=True
    )

    # Paginate: 20 combined records per page
    paginator = Paginator(all_receipts, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Group paginated records by creation date
    grouped_data = []
    for date, items in groupby(page_obj, key=lambda r: localtime(r.created_at).date()):
        items = list(items)
        category_totals = defaultdict(Decimal)
        daily_total = Decimal('0.00')

        for r in items:
            category = getattr(r, 'category', None) or getattr(r, 'purpose', 'unknown')
            category = category.lower().strip() if category else 'unknown'
            amount = r.amount or Decimal('0.00')
            category_totals[category] += amount
            daily_total += amount

        grouped_data.append({
            'date': date,
            'receipts': items,
            'category_totals': {
                'tithes': category_totals.get('tithes', Decimal('0.00')),
                'in_kind': category_totals.get('in_kind', Decimal('0.00')),
                'thanksgiving': category_totals.get('thanksgiving', Decimal('0.00')),
                'offering': category_totals.get('offering', Decimal('0.00')),
                'thanks': category_totals.get('thanks', Decimal('0.00')),
                'donation': category_totals.get('donation', Decimal('0.00')),
            },
            'grand_total': daily_total
        })

    context = {
        'grouped_by_date': grouped_data,
        'page_obj': page_obj,  # for pagination controls
    }
    return render(request, 'church/sunday_income_list.html', context)




@login_required
def delete_sunday_income_receipt(request, pk):
    receipt = get_object_or_404(SundayIncomeReceipt, pk=pk)  # <- removed is_deleted=False
    receipt.is_deleted = True
    receipt.save()
    return redirect('church:sunday_income_receipt_list')


class DeleteOnlineGivingView(LoginRequiredMixin, View):
    def post(self, request, pk):
        giving = get_object_or_404(OnlineGiving, pk=pk, church=request.user.churchadminprofile.church)
        giving.delete()
        return redirect('church:sunday_income_receipt_list') 



@login_required
def restore_sunday_income_receipt(request, pk):
    try:
        receipt = SundayIncomeReceipt.objects.get(pk=pk)
        if receipt.is_deleted:
            receipt.is_deleted = False
            receipt.save()
        # else it's already restored; no action needed
        return redirect('church:trashed_sunday_income_list')
    except SundayIncomeReceipt.DoesNotExist:
        raise Http404("Receipt not found.")


# Optional: View for deleted/trash receipts (admin only)
def deleted_receipts_list(request):
    deleted_receipts = SundayIncomeReceipt.objects.filter(is_deleted=True).order_by('-created_at')
    return render(request, 'church/deleted_receipts_list.html', {'deleted_receipts': deleted_receipts})




@login_required
def trashed_sunday_income_receipts(request):
    user = request.user

    if not user.is_church_admin:
        messages.error(request, "You are not authorized to view this page.")
        return redirect('church_admin_login')

    try:
        church = user.churchadminprofile.church
    except ChurchAdminProfile.DoesNotExist:
        messages.error(request, "Church admin profile not found.")
        return redirect('church_admin_login')

    receipts = SundayIncomeReceipt.objects.filter(church=church, is_deleted=True).order_by('-created_at')
    
    return render(request, 'church/trashed_sunday_income_list.html', {'receipts': receipts})


@login_required
def soft_delete_sunday_income_receipt(request, receipt_id):
    user = request.user

    # Ensure only church admins can proceed
    if not getattr(user, 'is_church_admin', False):
        return redirect('church_admin_login')

    try:
        church = user.churchadminprofile.church
    except ChurchAdminProfile.DoesNotExist:
        return redirect('church_admin_login')

    # Retrieve the receipt safely (and ensure it's not already deleted)
    receipt = get_object_or_404(
        SundayIncomeReceipt, id=receipt_id, church=church, is_deleted=False
    )

    # ✅ Perform soft delete by flagging
    receipt.is_deleted = True
    receipt.deleted_at = timezone.now()
    receipt.save()

    # Feedback to user
    messages.success(request, "Receipt moved to trash.")
    return redirect('church:sunday_income_receipt')


@login_required
def permanently_delete_sunday_income_receipt(request, receipt_id):
    user = request.user

    if not user.is_church_admin:
        messages.error(request, "You are not authorized to perform this action.")
        return redirect('church_admin_login')

    try:
        church = user.churchadminprofile.church
    except ChurchAdminProfile.DoesNotExist:
        messages.error(request, "Church admin profile not found.")
        return redirect('church_admin_login')

    try:
        receipt = SundayIncomeReceipt.objects.get(id=receipt_id, church=church, is_deleted=True)

        # ✅ Hard delete bypassing custom delete()
        models.Model.delete(receipt)

        messages.success(request, "Receipt permanently deleted.")
    except SundayIncomeReceipt.DoesNotExist:
        messages.error(request, "Receipt not found or already deleted.")

    receipts = SundayIncomeReceipt.objects.filter(church=church, is_deleted=True).order_by('-created_at')
    return render(request, 'church/trashed_sunday_income_list.html', {'receipts': receipts})



@login_required
def sunday_income_summary(request):
    church = request.user.churchadminprofile.church
    today = timezone.now().date()
    start_week = today - timedelta(days=today.weekday())
    start_month = today.replace(day=1)

    weekly_total = SundayIncomeReceipt.objects.filter(
        church=church,
        created_at__date__gte=start_week
    ).aggregate(total=Sum('amount'))['total'] or 0

    monthly_total = SundayIncomeReceipt.objects.filter(
        church=church,
        created_at__date__gte=start_month
    ).aggregate(total=Sum('amount'))['total'] or 0

    return render(request, 'church/sunday_income_summary.html', {
        'weekly_total': weekly_total,
        'monthly_total': monthly_total,
    })


def verify_receipt(request, receipt_number):
    try:
        receipt = SundayIncomeReceipt.objects.select_related('church').get(receipt_number=receipt_number)
        context = {'receipt': receipt, 'valid': True}
    except SundayIncomeReceipt.DoesNotExist:
        context = {'receipt_number': receipt_number, 'valid': False}
    
    return render(request, 'verify_receipt.html', context)


@staff_member_required
def regenerate_qr_admin_view(request, receipt_id):
    receipt = get_object_or_404(SundayIncomeReceipt, id=receipt_id)
    receipt.regenerate_qr_code()
    messages.success(request, f"QR Code regenerated for receipt {receipt.receipt_number}.")
    return redirect(request.META.get('HTTP_REFERER', '/admin/'))


@login_required
def download_filtered_receipts_pdf(request):
    form = SundayReceiptFilterForm(request.GET or None)

    receipts = SundayIncomeReceipt.objects.all()
    online_givings = OnlineGiving.objects.all()

    if form.is_valid():
        start_date = form.cleaned_data.get('start_date')
        end_date = form.cleaned_data.get('end_date')
        category = form.cleaned_data.get('category')

        if start_date:
            receipts = receipts.filter(created_at__date__gte=start_date)
            online_givings = online_givings.filter(created_at__date__gte=start_date)
        if end_date:
            receipts = receipts.filter(created_at__date__lte=end_date)
            online_givings = online_givings.filter(created_at__date__lte=end_date)
        if category:
            receipts = receipts.filter(category=category)
            online_givings = online_givings.filter(category=category)
    else:
        start_date = end_date = category = None

    # Filter by church of logged-in admin
    church = None
    if hasattr(request.user, 'churchadminprofile'):
        church = request.user.churchadminprofile.church
        receipts = receipts.filter(church=church)
        online_givings = online_givings.filter(church=church)

    # Normalize OnlineGiving
    for giving in online_givings:
        giving.name = giving.giver_name
        giving.receiver_name = giving.church.church_name
        giving.source_type = 'online'
        giving.phone_number = getattr(giving, 'phone_number', '')

    # Normalize SundayIncomeReceipt
    for receipt in receipts:
        receipt.name = getattr(receipt, 'name', "—")
        receipt.receiver_name = getattr(receipt, 'receiver_name', "—")
        receipt.source_type = 'receipt'
        receipt.phone_number = getattr(receipt, 'phone_number', '')

    # Combine and sort all
    all_receipts = sorted(
        list(receipts) + list(online_givings),
        key=lambda r: localtime(r.created_at).date()
    )

    # Group by date
    grouped_by_date = []
    grand_total = Decimal('0.00')
    for date, entries in groupby(all_receipts, key=lambda r: localtime(r.created_at).date()):
        items = list(entries)
        date_total = sum((item.amount or Decimal('0.00')) for item in items)
        grand_total += date_total
        grouped_by_date.append({
            'date': date,
            'items': items,
            'date_total': date_total
        })

    # Totals by category
    category_totals = {
        'tithes': (
            receipts.filter(category='tithes').aggregate(total=Sum('amount'))['total'] or 0
        ) + (
            online_givings.filter(category='tithes').aggregate(total=Sum('amount'))['total'] or 0
        ),
        'in_kind': (
            receipts.filter(category='in_kind').aggregate(total=Sum('amount'))['total'] or 0
        ) + (
            online_givings.filter(category='in_kind').aggregate(total=Sum('amount'))['total'] or 0
        ),
        'thanksgiving': (
            receipts.filter(category='thanksgiving').aggregate(total=Sum('amount'))['total'] or 0
        ) + (
            online_givings.filter(category='thanksgiving').aggregate(total=Sum('amount'))['total'] or 0
        ),
        'offering': (
            receipts.filter(category='offering').aggregate(total=Sum('amount'))['total'] or 0
        ) + (
            online_givings.filter(category='offering').aggregate(total=Sum('amount'))['total'] or 0
        ),
    }

    logo_url = request.build_absolute_uri(church.profile_picture.url) if church and church.profile_picture else None

    template = get_template('church/sunday_receipts_pdf.html')
    html = template.render({
        'grouped_by_date': grouped_by_date,
        'category_totals': category_totals,
        'grand_total': grand_total,
        'form': form,
        'start_date': start_date,
        'end_date': end_date,
        'category': category,
        'church': church,
        'church_logo_url': logo_url,
        'now': now(),
    })

    pdf_file = BytesIO()
    HTML(string=html, base_url=request.build_absolute_uri('/')).write_pdf(pdf_file)
    pdf_file.seek(0)

    response = HttpResponse(pdf_file.read(), content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="sunday_receipts.pdf"'
    return response




@method_decorator(login_required, name='dispatch')
class OnlineGivingView(View):
    def get(self, request):
        form = OnlineGivingForm()
        receipt_id = request.GET.get('receipt_id')
        return render(request, 'church/online_giving.html', {
            'form': form,
            'receipt_id': receipt_id
        })

    def post(self, request):
        form = OnlineGivingForm(request.POST)
        if form.is_valid():
            online_giving = form.save(commit=False)
            online_giving.member = request.user.member
            online_giving.church = request.user.member.church
            online_giving.save()

            return redirect(reverse('church:online_giving') + f'?receipt_id={online_giving.id}')

        return render(request, 'church/online_giving.html', {'form': form})



@method_decorator(login_required, name='dispatch')
class SMSLogListView(ListView):
    model = SMSLog
    template_name = 'church/sms_log_list.html'
    context_object_name = 'logs'

    def get_queryset(self):
        user = self.request.user
        queryset = SMSLog.objects.none()
        if hasattr(user, 'churchadminprofile'):
            queryset = SMSLog.objects.filter(church=user.churchadminprofile.church)

            q = self.request.GET.get('q')
            if q:
                queryset = queryset.filter(
                    Q(phone_number__icontains=q) |
                    Q(message__icontains=q) |
                    Q(sent_at__icontains=q)
                )
            queryset = queryset.order_by('-sent_at')
        return queryset



@method_decorator(login_required, name='dispatch')
class ResendSMSView(View):
    def post(self, request, pk):
        sms_log = get_object_or_404(SMSLog, pk=pk, church=request.user.churchadminprofile.church)
        response = send_sms(sms_log.phone_number, sms_log.message, church=sms_log.church)
        if response.get("return"):
            sms_log.sent_at = now()
            sms_log.success = True
            sms_log.save()
            messages.success(request, f"SMS resent successfully to {sms_log.phone_number}.")
        else:
            messages.error(request, "Failed to resend SMS. Check Fast2SMS configuration or logs.")
        return redirect('sms_logs')



class BibleReadingPlanListView(LoginRequiredMixin, ListView):
    model = BibleReadingPlan
    template_name = 'church/reading_plan_list.html'
    context_object_name = 'plans'

    def get_queryset(self):
        return BibleReadingPlan.objects.filter(
            church=self.request.user.churchadminprofile.church
        ).order_by('-reading_date')


class BibleReadingPlanCreateView(LoginRequiredMixin, CreateView):
    model = BibleReadingPlan
    form_class = BibleReadingPlanForm
    template_name = 'church/reading_plan_form.html'
    success_url = reverse_lazy('church:bible_reading_list')

    def form_valid(self, form):
        verse_text = self.request.POST.get('verse_text', '').strip()
        commentary_text = self.request.POST.get('commentary_text', '').strip()

        # Clean verse_text if it's accidentally JSON
        try:
            verse_json = json.loads(verse_text)
            verse_text = verse_json.get('text', verse_text)
        except json.JSONDecodeError:
            pass

        # Clean commentary_text if it's accidentally JSON
        if commentary_text.startswith('{') and commentary_text.endswith('}'):
            try:
                comm_json = json.loads(commentary_text)
                commentary_text = comm_json.get('text', commentary_text)
            except json.JSONDecodeError:
                pass

        content = {
            'reference': form.cleaned_data['scripture_reference'],
            'verses': [{'verse': '', 'text': verse_text}] if verse_text else [],
            'commentary': commentary_text
        }

        form.instance.content = json.dumps(content)
        form.instance.church = self.request.user.churchadminprofile.church
        return super().form_valid(form)


class BibleReadingPlanUpdateView(LoginRequiredMixin, UpdateView):
    model = BibleReadingPlan
    form_class = BibleReadingPlanForm
    template_name = 'church/reading_plan_form.html'
    success_url = reverse_lazy('church:bible_reading_list')

    def get_initial(self):
        initial = super().get_initial()
        try:
            content = json.loads(self.object.content or '{}')
            # Extract first verse text if available
            if 'verses' in content and content['verses']:
                initial['verse_text'] = content['verses'][0].get('text', '')
            # Extract commentary
            initial['commentary_text'] = content.get('commentary', '')
        except json.JSONDecodeError:
            initial['verse_text'] = ''
            initial['commentary_text'] = ''
        return initial

    def form_valid(self, form):
        verse_text = self.request.POST.get('verse_text', '').strip()
        commentary_text = self.request.POST.get('commentary_text', '').strip()

        try:
            verse_json = json.loads(verse_text)
            verse_text = verse_json.get('text', verse_text)
        except json.JSONDecodeError:
            pass

        if commentary_text.startswith('{') and commentary_text.endswith('}'):
            try:
                comm_json = json.loads(commentary_text)
                commentary_text = comm_json.get('text', commentary_text)
            except json.JSONDecodeError:
                pass

        content = {
            'reference': form.cleaned_data['scripture_reference'],
            'verses': [{'verse': '', 'text': verse_text}] if verse_text else [],
            'commentary': commentary_text
        }

        form.instance.content = json.dumps(content)
        return super().form_valid(form)



class BibleReadingPlanDeleteView(LoginRequiredMixin, DeleteView):
    model = BibleReadingPlan
    template_name = 'church/reading_plan_confirm_delete.html'
    success_url = reverse_lazy('church:bible_reading_list')

    def get_queryset(self):
        return BibleReadingPlan.objects.filter(
            church=_get_user_church_generic(self.request.user)
        )



class MemberBibleReadingTodayView(LoginRequiredMixin, TemplateView):
    template_name = 'church/reading_today.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        member = getattr(self.request.user, 'member', None)

        if not member:
            context.update({
                'reading': None,
                'parsed_verses': [],
                'commentary': '',
                'error': "Member profile not found."
            })
            return context

        today = now().date()
        reading = BibleReadingPlan.objects.filter(
            church=member.church,
            reading_date=today
        ).first()

        if reading:
            try:
                content = json.loads(reading.content or '{}')
                context.update({
                    'reading': reading,
                    'parsed_verses': content.get('verses', []),
                    'commentary': content.get('commentary', ''),
                    'error': None
                })
            except json.JSONDecodeError:
                context.update({
                    'reading': reading,
                    'parsed_verses': [],
                    'commentary': '',
                    'error': "❌ Failed to parse reading content."
                })
        else:
            context.update({
                'reading': None,
                'parsed_verses': [],
                'commentary': '',
                'error': "📌 No devotion available today."
            })

        return context


class MemberBibleReadingArchiveView(LoginRequiredMixin, ListView):
    model = BibleReadingPlan
    template_name = 'church/reading_archive.html'
    context_object_name = 'plans'

    def get_queryset(self):
        church = self.request.user.member.church
        queryset = BibleReadingPlan.objects.filter(church=church).order_by('-reading_date')

        for plan in queryset:
            try:
                parsed = json.loads(plan.content or '{}')
                plan.parsed_verses_data = parsed.get('verses', [])
            except Exception:
                plan.parsed_verses_data = []
        return queryset


class MarkReadingAsDoneView(LoginRequiredMixin, View):
    def post(self, request, *args, **kwargs):
        reading_id = request.POST.get('reading_id')

        if not reading_id:
            return JsonResponse({'status': 'error', 'message': 'Missing reading ID'})

        try:
            reading = BibleReadingPlan.objects.get(pk=reading_id)
        except BibleReadingPlan.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Reading not found'})

        try:
            member = Member.objects.get(user=request.user)
        except Member.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Member not found'})

        reading_log, created = MemberReadingLog.objects.get_or_create(
            member=member,
            reading_plan=reading
        )

        if created:
            ActivityLog.objects.create(
                member=request.user,  # Assuming ActivityLog.member = ForeignKey(User)
                activity_type="Devotional",
                description=f"Marked devotional for {reading.reading_date} as read"
            )
            return JsonResponse({'status': 'marked', 'message': '✅ Devotional marked as read!'})
        else:
            return JsonResponse({'status': 'already_marked', 'message': 'ℹ️ You already marked this as read.'})



class MemberReadingProgressView(LoginRequiredMixin, TemplateView):
    template_name = 'church/reading_progress.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        member = Member.objects.get(user=self.request.user)
        church = member.church
        today = now().date()
        start_of_month = today.replace(day=1)

        monthly_readings = BibleReadingPlan.objects.filter(
            church=church,
            reading_date__range=(start_of_month, today)
        ).count()

        monthly_completed = MemberReadingLog.objects.filter(
            member=member,
            reading_plan__reading_date__range=(start_of_month, today)
        ).count()

        progress_percent = (monthly_completed / monthly_readings * 100) if monthly_readings else 0

        context.update({
            'monthly_readings': monthly_readings,
            'monthly_completed': monthly_completed,
            'progress_percent': round(progress_percent, 2)
        })
        return context



class MemberActivityLogView(LoginRequiredMixin, ListView):
    model = ActivityLog
    template_name = 'church/activity_log.html'  # Update this if needed
    context_object_name = 'logs'

    def get_queryset(self):
        return ActivityLog.objects.filter(
            member=self.request.user
        ).order_by('-timestamp')


class PrayerRequestListView(LoginRequiredMixin, ListView):
    model = PrayerRequest
    template_name = "church/prayer_requests.html"
    context_object_name = "prayers"

    def get_queryset(self):
        return (
            PrayerRequest.objects.active()        # <- NEW
            .filter(member__member__church=self.request.user.member.church)
        )


class PrayerRequestCreateView(LoginRequiredMixin, CreateView):
    model = PrayerRequest
    form_class = PrayerRequestForm
    template_name = 'church/add_prayer_request.html'
    success_url = reverse_lazy('church:prayer_request_list')

    def form_valid(self, form):
        form.instance.member = self.request.user
        return super().form_valid(form)


@login_required
def generate_society_members_pdf(request, society_name):
    user = request.user
    church = getattr(user, 'churchadminprofile', None)
    if church:
        church = church.church
    elif user.is_superuser:
        church_id = request.GET.get('church_id')
        if not church_id:
            return HttpResponse("Superuser must provide ?church_id=", status=400)
        church = Church.objects.filter(id=church_id).first()

    if not church:
        return HttpResponseNotFound("Church not found.")

    # ✅ Filter by category name, which represents the society
    members = Member.objects.filter(
        church=church,
        category__name__iexact=society_name
    ).order_by('first_name', 'last_name')

    context = {
        'church_name': church.church_name,
        'church_image_url': church.profile_picture.url if church.profile_picture else None,
        'society_name': society_name,
        'members': members,
        'now': timezone.now(),
    }

    html_string = get_template('church/society_members_pdf.html').render(context)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{society_name.lower()}_members.pdf"'

    fd, path = tempfile.mkstemp(suffix=".pdf")
    try:
        HTML(string=html_string, base_url=request.build_absolute_uri('/')).write_pdf(path)
        with open(path, 'rb') as pdf_file:
            response.write(pdf_file.read())
    finally:
        os.close(fd)
        os.remove(path)

    return response



def submit_advertisement(request):
    if request.method == 'POST':
        form = AdvertisementForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()  # Not approved yet
            return render(request, 'ads/thank_you.html')
    else:
        form = AdvertisementForm()
    return render(request, 'ads/submit_ad.html', {'form': form})


def display_approved_ads(request):
    ads = Advertisement.objects.filter(is_approved=True)
    return render(request, 'ads/show_ads.html', {'ads': ads})



def home(request):
    testimonial_list = Testimonial.objects.order_by('-created_at')
    paginator = Paginator(testimonial_list, 3)  # Show 3 testimonials per page
    page_number = request.GET.get('page')
    testimonials = paginator.get_page(page_number)

    chairman_msg = ChairmanMessage.objects.first()
    ads = Advertisement.objects.all()
    gallery_images = ChurchGalleryImage.objects.all()

    return render(
        request,
        'church/homepage.html',
        {
            'testimonials': testimonials,
            'chairman_msg': chairman_msg,
            'ads': ads,
            'gallery_images': gallery_images,
        }
    )




def license_entry_view(request):
    if request.method == "POST":
        form = LicenseValidationForm(request.POST)
        if form.is_valid():
            # Save the key in session
            request.session['validated_license_key'] = form.cleaned_data['license_key']
            return redirect('church:register')  # registration view
    else:
        form = LicenseValidationForm()
    return render(request, 'church/license_entry.html', {'form': form})





class PaymentInstructionsView(TemplateView):
    template_name = "church/make_payment.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Grab the most‑recent QR/UPI record, if any
        qr_rec = PaymentQRCode.objects.order_by("-uploaded_at").first()

        # ---- UPI ID --------------------------------------------------------
        context["upi_id"] = (
            qr_rec.upi_id if qr_rec and qr_rec.upi_id else settings.PAYMENT_UPI_ID
        )

        # ---- QR image URL --------------------------------------------------
        # We keep the original key name `qr_path` so your template stays unchanged.
        context["qr_path"] = (
            qr_rec.qr_image.url if qr_rec and qr_rec.qr_image else settings.PAYMENT_QR_PATH
        )

        # ---- Instructions (unchanged) -------------------------------------
        context["instructions"] = settings.PAYMENT_INSTRUCTIONS
        return context



def _get_user_church_generic(user):
    if hasattr(user, "member"):
        return user.member.church
    if hasattr(user, "volunteer"):
        return user.volunteer.church
    if hasattr(user, "churchadminprofile"):
        return user.churchadminprofile.church
    raise PermissionDenied("User is not linked to a church.")



@login_required
@require_GET
def fetch_member_chat(request):
    """
    Return up to 50 messages for the member chat, with *unread*
    (i.e. sent after the user’s last fetch) first, followed by
    already‑seen messages – both blocks ordered oldest → newest.
    """
    church = _get_user_church_generic(request.user)

    # ── 1. last‑seen timestamp stored in the session ────────────────
    last_seen_iso = request.session.get("member_chat_last_seen")          # "2025‑07‑15T09:21:00.000000Z"
    last_seen     = None
    if last_seen_iso:
        try:
            last_seen = timezone.datetime.fromisoformat(last_seen_iso)
            if timezone.is_naive(last_seen):                # make aware if needed
                last_seen = timezone.make_aware(last_seen, timezone.utc)
        except ValueError:
            last_seen = None                                 # ignore corrupt value

    # ── 2. base queryset ─────────────────────────────────────────────
    qs_base = (
        MemberChatMessage.objects
        .filter(church=church, is_deleted=False)
        .select_related("sender")
    )

    # ── 3. unread first, then read (both oldest → newest) ───────────
    if last_seen:
        unread_qs = qs_base.filter(sent_at__gt=last_seen).order_by("sent_at")
        read_qs   = qs_base.filter(sent_at__lte=last_seen).order_by("sent_at")
        msgs      = list(unread_qs) + list(read_qs)
    else:
        # first visit → everything is “unread”
        msgs = list(qs_base.order_by("sent_at"))

    # keep only the last 50 messages overall
    msgs = msgs[-50:]

    # ── 4. store *now* as the new “last seen” marker ────────────────
    request.session["member_chat_last_seen"] = timezone.now().isoformat()

    # ── 5. permission helpers for delete icon ───────────────────────
    is_volunteer = hasattr(request.user, "volunteer")
    is_mod       = request.user.has_perm("church.moderate_message")

    # ── 6. JSON payload ──────────────────────────────────────────────
    return JsonResponse({
        "messages": [
            {
                "id"        : m.id,
                "sender"    : m.sender.first_name or m.sender.email,
                "content"   : m.content,
                "ago"       : f"{timesince(m.sent_at)} ago",
                "can_delete": (
                    is_mod or
                    is_volunteer or
                    m.sender_id == request.user.id     # ← members delete their own
                ),
            }
            for m in msgs
        ]
    })





@login_required
@require_POST
def post_member_chat(request):
    content = request.POST.get("content", "").strip()
    if not content:
        return JsonResponse({"ok": False, "error": "Empty message"}, status=400)

    church = _get_user_church_generic(request.user)

    MemberChatMessage.objects.create(
        sender  = request.user,
        content = content[:1000],
        church  = church
    )

    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        return JsonResponse({"ok": True})
    return HttpResponseRedirect(request.META.get("HTTP_REFERER", "/"))



# ── Volunteer List page ────────────────────────────────
@method_decorator(permission_required("church.moderate_message", raise_exception=True), name='dispatch')
class MemberChatModerationListView(ListView):
    template_name = "volunteer/chat_moderation_list.html"
    paginate_by   = 50

    def get_queryset(self):
        # Only show messages from the volunteer’s church
        return (
            MemberChatMessage.objects
            .filter(
                church      = self.request.user.volunteer.church,
                is_deleted  = False
            )
            .select_related("sender", "church")
            .order_by("-sent_at")
        )


@require_POST
@login_required
def delete_member_chat_message(request, pk):
    """
    Volunteers may delete **any** message.
    Members  may delete **only their own**.
    Moderators (perm `church.moderate_message`) may delete any message.
    """
    msg    = get_object_or_404(MemberChatMessage, pk=pk, is_deleted=False)
    church = _get_user_church_generic(request.user)

    if msg.church != church:
        return JsonResponse({"error": "Unauthorized"}, status=403)

    # ── permission checks ──────────────────────────────────────────────
    is_volunteer = hasattr(request.user, "volunteer")
    is_mod       = request.user.has_perm("church.moderate_message")
    is_sender    = msg.sender_id == request.user.id

    if not (is_volunteer or is_mod or is_sender):
        return JsonResponse({"error": "Unauthorized"}, status=403)
    # ------------------------------------------------------------------

    msg.is_deleted = True
    msg.deleted_by = request.user
    msg.save(update_fields=["is_deleted", "deleted_by"])

    wants_json = (
        request.headers.get("x-requested-with") == "XMLHttpRequest" or
        request.headers.get("accept", "").startswith("application/json")
    )

    if wants_json:
        return JsonResponse({"status": "ok", "id": pk})

    return HttpResponseRedirect(request.META.get("HTTP_REFERER", "/"))


@method_decorator(login_required, name="dispatch")
class ChurchOnlineGivingView(View):
    """Dashboard screen where the church admin enters UPI ID & uploads QR image."""
    template_name = "church/church_online_giving.html"

    def _get_church(self, user):
        # Your Church model uses OneToOne → related_name="church_admin"
        return getattr(user, "church_admin", None)

    def get(self, request):
        church = self._get_church(request.user)
        if not church:
            return render(request, "403.html")
        form = ChurchOnlineGivingForm(instance=church)
        return render(request, self.template_name, {"form": form})

    def post(self, request):
        church = self._get_church(request.user)
        if not church:
            return render(request, "403.html")
        form = ChurchOnlineGivingForm(request.POST, request.FILES, instance=church)
        if form.is_valid():
            form.save()
            messages.success(request, "Online‑giving details saved.")
            # ✅ include namespace in redirect:
            return redirect("church:church_admin_online_giving")
        return render(request, self.template_name, {"form": form})





# Define role priority
ROLE_PRIORITY = {
    "pastor": 1,
    "associate_pastor": 2,
    "head_deacon": 3,
    "clerk": 4,
    "secretary": 4,
    "deacon": 5,
    "deaconess": 6,
    "song leader": 7,
    "choir master": 7,
    "pianist": 8,
    "auditor": 9,
    "sound system technician": 10,
    "chowkidar": 11,
    "projector in charge": 12,
}

@login_required
def staff_list(request):
    church = request.user.churchadminprofile.church
    staff_members_qs = ChurchStaffMember.objects.filter(church=church, is_active=True)

    # Sort by role priority
    staff_members = sorted(
        staff_members_qs,
        key=lambda member: ROLE_PRIORITY.get(member.role.lower(), 99)  # Default low priority if missing
    )

    return render(request, 'church/staff_list.html', {'staff_members': staff_members})


@login_required
def add_staff(request):
    church = request.user.churchadminprofile.church
    if request.method == 'POST':
        form = ChurchStaffMemberForm(request.POST, request.FILES)
        if form.is_valid():
            staff = form.save(commit=False)
            staff.church = church
            staff.save()
            messages.success(request, "Staff member added successfully.")
            return redirect('church:staff_list')
    else:
        form = ChurchStaffMemberForm()

    return render(request, 'church/add_staff_member.html', {'form': form})


@login_required
def edit_staff(request, staff_id):
    church = request.user.churchadminprofile.church
    staff = get_object_or_404(ChurchStaffMember, id=staff_id, church=church)
    if request.method == 'POST':
        form = ChurchStaffMemberForm(request.POST, request.FILES, instance=staff)
        if form.is_valid():
            form.save()
            messages.success(request, "Staff member updated.")
            return redirect('church:staff_list')
    else:
        form = ChurchStaffMemberForm(instance=staff)
    return render(request, 'church/edit_staff.html', {'form': form, 'staff': staff})


@login_required
@require_POST
def delete_staff(request, staff_id):
    church = request.user.churchadminprofile.church
    staff = get_object_or_404(ChurchStaffMember, id=staff_id, church=church)
    staff.delete()
    messages.success(request, "Staff member deleted successfully.")
    return redirect('church:staff_list')


@login_required
def add_staff_payment(request):
    try:
        church = request.user.churchadminprofile.church
    except AttributeError:
        messages.error(request, "Your church admin profile is not set up correctly.")
        return redirect('church:church_admin_dashboard')

    if request.method == 'POST':
        form = StaffPaymentForm(request.POST, church=church)
        if form.is_valid():
            payment = form.save(commit=False)
            payment.church = church  # ✅ Assign church
            payment.save()
            messages.success(request, "✅ Staff payment recorded successfully.")
            return redirect('church:staff_salary_table')
        else:
            messages.error(request, "⚠️ Please correct the form errors.")
    else:
        form = StaffPaymentForm(church=church)

    return render(request, 'church/add_staff_payment.html', {'form': form})


@login_required
def staff_payment_list(request):
    church = request.user.churchadminprofile.church
    payments = StaffPayment.objects.filter(
        church=church
    ).select_related('staff').order_by('-paid_on')
    return render(request, 'church/staff_payment_list.html', {'payments': payments})




# Define role priorities
ROLE_PRIORITY = {
    "pastor": 1,
    "associate_pastor": 2,
    "head_deacon": 3,
    "clerk": 4,
    "secretary": 4,
    "deacon": 5,
    "deaconess": 6,
    "choir_master": 7,
    "song leader": 7,
    "pianist": 8,
    "auditor": 9,
    "sound_technician": 10,
    "chowkidar": 11,
    "projector_operator": 12,
}


class StaffSalaryTableView(TemplateView):
    template_name = 'church/staff_salary_table.html'

    ROLE_PRIORITY = {
        "pastor": 1,
        "associate_pastor": 2,
        "head_deacon": 3,
        "clerk": 4,
        "secretary": 4,
        "deacon": 5,
        "deaconess": 6,
        "choir_master": 7,
        "song leader": 7,
        "pianist": 8,
        "auditor": 9,
        "sound_technician": 10,
        "chowkidar": 11,
        "projector_operator": 12,
    }

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        church = user.churchadminprofile.church

        selected_year = int(self.request.GET.get('year', datetime.now().year))
        year_range = list(range(2020, datetime.now().year + 3))  # for dropdown

        staff_qs = ChurchStaffMember.objects.filter(church=church, is_active=True)
        staff_qs = sorted(staff_qs, key=lambda m: self.ROLE_PRIORITY.get(m.role.lower(), 999))

        all_payments = StaffPayment.objects.filter(
            church=church,
            payment_month__year=selected_year
        ).select_related('staff')

        payment_map = defaultdict(lambda: {m: None for m in range(1, 13)})
        for p in all_payments:
            payment_map[p.staff.id][p.payment_month.month] = p

        staff_data = []
        for staff in staff_qs:
            row = {
                'staff': staff,
                'payments': payment_map.get(staff.id, {m: None for m in range(1, 13)})
            }
            staff_data.append(row)

        monthly_totals = {m: 0 for m in range(1, 13)}
        for row in staff_data:
            for m, payment in row['payments'].items():
                if payment:
                    monthly_totals[m] += payment.amount

        grand_total = sum(monthly_totals.values())

        context.update({
            'staff_data': staff_data,
            'month_range': range(1, 13),
            'month_names': [month_name[m] for m in range(1, 13)],
            'year': selected_year,
            'year_range': year_range,
            'monthly_totals': monthly_totals,
            'grand_total': grand_total
        })

        return context


@login_required
def edit_staff_payment(request, pk):
    payment = get_object_or_404(StaffPayment, pk=pk)
    church = payment.staff.church  # Important!

    if request.method == 'POST':
        form = StaffPaymentForm(request.POST, instance=payment, church=church)
        if form.is_valid():
            form.save()
            return redirect('church:staff_salary_table')  # or use reverse()
    else:
        form = StaffPaymentForm(instance=payment, church=church)

    return render(request, 'church/edit_staff_payment.html', {
        'form': form,
        'payment': payment
    })



def delete_staff_payment(request, pk):
    try:
        payment = StaffPayment.objects.get(pk=pk)
    except StaffPayment.DoesNotExist:
        return redirect(reverse('church:staff_salary_table'))  # fallback if already deleted

    if request.method == 'POST':
        payment.delete()
        return redirect(reverse('church:staff_salary_table'))

    return render(request, 'church/delete_staff_payment.html', {'payment': payment})



ROLE_PRIORITY = {
    "pastor": 1,
    "associate_pastor": 2,
    "head_deacon": 3,
    "clerk": 4,
    "secretary": 4,
    "deacon": 5,
    "deaconess": 6,
    "choir_master": 7,
    "song leader": 7,
    "pianist": 8,
    "auditor": 9,
    "sound_technician": 10,
    "chowkidar": 11,
    "projector_operator": 12,
}

def get_role_priority(role):
    return ROLE_PRIORITY.get(role.lower().replace(" ", "_"), 999)

@login_required
@login_required
def staff_salary_pdf_view(request):
    year = int(request.GET.get("year", now().year))
    church = request.user.churchadminprofile.church

    staff_members = ChurchStaffMember.objects.filter(church=church)
    month_range = range(1, 13)

    staff_data = []
    for staff in staff_members:
        payments = {
            p.payment_month.month: p
            for p in StaffPayment.objects.filter(
                staff=staff, payment_month__year=year
            )
        }
        staff_data.append({
            "staff": staff,
            "payments": payments
        })

    monthly_totals = {}
    grand_total = 0
    for month in month_range:
        total = sum(
            payments.get(month).amount
            for row in staff_data
            if (payments := row["payments"]) and month in payments
        )
        monthly_totals[month] = total
        grand_total += total

    logo_url = ""
    if church.profile_picture:
        logo_url = request.build_absolute_uri(church.profile_picture.url)

    template = get_template("church/staff_salary_pdf.html")
    html = template.render({
        "staff_data": staff_data,
        "year": year,
        "month_range": month_range,
        "monthly_totals": monthly_totals,
        "grand_total": grand_total,
        "church_name": church.church_name,
        "church_logo": logo_url,
    })

    pdf_file = HTML(string=html).write_pdf(stylesheets=[
        CSS(string='@page { size: A4 landscape; margin: 1cm; }')
    ])

    response = HttpResponse(pdf_file, content_type="application/pdf")
    response["Content-Disposition"] = f"filename=Staff_Salary_Table_{year}.pdf"
    return response





